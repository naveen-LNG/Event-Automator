import sys
import os
import openpyxl
import json
import re

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
    except IndexError:
        print("Error: Missing design folder path")
        sys.exit(1)

    bp_path = os.path.join(design_path, "festivals_bp.xlsx")
    
    data = {
        "last_cycle_id": 0,
        "last_recommend_id": 0,
        "last_item_id": 0,
        "last_scheme_id": 0,
        "last_holiday_id_str": "",
        "last_suffixes": [],
        "last_title_set": 1 # Default
    }

    try:
        wb = openpyxl.load_workbook(bp_path, data_only=True)
        
        # 1. festivals_bp sheet
        ws_bp = wb['festivals_bp']
        last_bp_row = 1
        for r in range(ws_bp.max_row, 1, -1):
            if ws_bp.cell(row=r, column=1).value:
                last_bp_row = r
                break
        
        data["last_cycle_id"] = int(ws_bp.cell(row=last_bp_row, column=1).value or 0)
        data["last_recommend_id"] = int(ws_bp.cell(row=last_bp_row, column=9).value or 18400)
        last_title = str(ws_bp.cell(row=last_bp_row, column=5).value or "")
        data["last_title_set"] = 2 if "_2" in last_title else 1

        # 2. 扩展表 (Extended Table) - Find last 4-group
        ws_ext = wb['扩展表']
        last_ext_row = 1
        for r in range(ws_ext.max_row, 1, -1):
            if ws_ext.cell(row=r, column=1).value:
                last_ext_row = r
                break
        
        data["last_item_id"] = int(ws_ext.cell(row=last_ext_row, column=5).value or 3310)
        
        # Detect suffixes from the last 4 rows
        start_group = max(2, last_ext_row - 3)
        suffixes = []
        holiday_id_str = ""
        for r in range(start_group, last_ext_row + 1):
            alias = str(ws_ext.cell(row=r, column=2).value or "")
            match = re.match(r'(\d{6})(.*)', alias)
            if match:
                holiday_id_str = match.group(1)
                suffixes.append(match.group(2))
        
        data["last_holiday_id_str"] = holiday_id_str
        data["last_suffixes"] = suffixes

        # 3. chest_plan
        ws_cp = wb['chest_plan']
        last_cp_row = 1
        for r in range(ws_cp.max_row, 1, -1):
            if ws_cp.cell(row=r, column=1).value:
                last_cp_row = r
                break
        data["last_scheme_id"] = int(ws_cp.cell(row=last_cp_row, column=1).value or 1000)

        print(json.dumps(data))

    except Exception as e:
        print(f"Python Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
