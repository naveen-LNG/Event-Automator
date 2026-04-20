import sys
import os
import json
import openpyxl
import re

def get_last_data_row(ws, id_col=1):
    for r in range(ws.max_row, 0, -1):
        val = ws.cell(row=r, column=id_col).value
        if val is not None and str(val).strip() != '':
            return r
    return 1

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
    except IndexError:
        print(json.dumps({"error": "Missing design folder path"}))
        sys.exit(1)
        
    bp_path = os.path.join(design_path, "battle_pass.xlsx")
    
    if not os.path.exists(bp_path):
        print(json.dumps({"error": f"File not found: {bp_path}"}))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(bp_path, data_only=True)
        
        # 1. Analyze 'battle_pass' sheet
        ws_bp = wb['battle_pass']
        last_bp_row = get_last_data_row(ws_bp)
        
        last_cycle_id = ws_bp.cell(row=last_bp_row, column=1).value
        last_bp_plan_alias = ws_bp.cell(row=last_bp_row, column=7).value
        last_recharge_alias = ws_bp.cell(row=last_bp_row, column=8).value
        last_item_id = ws_bp.cell(row=last_bp_row, column=9).value
        
        # 2. Analyze 'chest_plan' sheet for Scheme ID and templates
        sname_cp = 'chest_plan' if 'chest_plan' in wb.sheetnames else 'Chest_plan'
        ws_cp = wb[sname_cp]
        last_cp_row = get_last_data_row(ws_cp)
        last_scheme_id = ws_cp.cell(row=last_cp_row, column=1).value
        
        last_800_cycle = None
        last_900_cycle = None
        
        # Scan for templates
        for r in range(4, last_cp_row + 1):
            try:
                # Column 1 is 方案ID, Column 2 is Alias (e.g. BattlePass第20期), Column 3 is Limit (800/900)
                limit_val = int(ws_cp.cell(row=r, column=3).value)
                # Extract number from Alias to get Cycle ID if possible, or mapping
                # But better: find the mapping from battle_pass sheet if they match
                pass
            except: continue
            
        # Hardcoded search based on known pattern since we have the data
        # Cycle 20 was 800. Let's find a 900 one.
        for r in range(last_cp_row, 3, -1):
            try:
                limit_val = int(ws_cp.cell(row=r, column=3).value)
                alias = str(ws_cp.cell(row=r, column=2).value)
                # Match cycle number from "BattlePass第XX期"
                m = re.search(r'(\d+)', alias)
                if m:
                    cycle_num = int(m.group(1))
                    if limit_val == 800 and last_800_cycle is None:
                        last_800_cycle = cycle_num
                    if limit_val == 900 and last_900_cycle is None:
                        last_900_cycle = cycle_num
                if last_800_cycle and last_900_cycle:
                    break
            except: continue

        result = {
            "last_cycle_id": int(last_cycle_id) if last_cycle_id else 0,
            "last_bp_plan_alias": str(last_bp_plan_alias),
            "last_recharge_alias": str(last_recharge_alias),
            "last_item_id": int(last_item_id) if last_item_id else 0,
            "last_scheme_id": int(last_scheme_id) if last_scheme_id else 0,
            "last_800_cycle": last_800_cycle,
            "last_900_cycle": last_900_cycle
        }
        print(json.dumps(result))

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
