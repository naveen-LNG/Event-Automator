import os
import sys
import shutil
import json
import traceback
import re

UNDO_LOG_FILE = "undo_log.json"

def log(msg):
    try:
        print(f"[PythonBackend] {msg}", flush=True)
    except UnicodeEncodeError:
        safe_msg = str(msg).encode('ascii', errors='replace').decode('ascii')
        print(f"[PythonBackend] {safe_msg}", flush=True)

def load_undo_log(root_path):
    log_path = os.path.join(root_path, UNDO_LOG_FILE)
    if os.path.exists(log_path):
        with open(log_path, 'r') as f:
            return json.load(f)
    return {}

def save_undo_log(root_path, undo_data):
    log_path = os.path.join(root_path, UNDO_LOG_FILE)
    with open(log_path, 'w') as f:
        json.dump(undo_data, f, indent=4)

def record_file_creation(root_path, step_index, file_path):
    undo_data = load_undo_log(root_path)
    step_key = str(step_index)
    
    if step_key not in undo_data:
        undo_data[step_key] = {"created": [], "modified": []}
        
    if file_path not in undo_data[step_key]["created"]:
        undo_data[step_key]["created"].append(file_path)
        
    save_undo_log(root_path, undo_data)

def revert_step(root_path, step_index):
    undo_data = load_undo_log(root_path)
    step_key = str(step_index)
    
    if step_key not in undo_data:
        log(f"No undo data found for Step {step_index}")
        return True
        
    files_created = undo_data[step_key].get("created", [])
    import stat
    
    for f in files_created:
        if os.path.exists(f):
            if f.endswith(".bak"):
                # It's a backup file, restore it
                orig_file = f[:-4]
                if os.path.exists(orig_file):
                    # Remove read-only from original before overwrite
                    os.chmod(orig_file, stat.S_IWRITE | stat.S_IREAD)
                shutil.copy2(f, orig_file)
                os.chmod(f, stat.S_IWRITE)
                os.remove(f)
                log(f"Reverted (Restored from bak): {os.path.basename(orig_file)}")
            else:
                try:
                    if os.path.isdir(f):
                        shutil.rmtree(f)
                        log(f"Reverted (Deleted Dir): {os.path.basename(f)}")
                    else:
                        os.chmod(f, stat.S_IWRITE)
                        os.remove(f)
                        log(f"Reverted (Deleted File): {os.path.basename(f)}")
                except Exception as e:
                    log(f"Failed to revert {f}: {e}")
                
    # Remove entry
    del undo_data[step_key]
    save_undo_log(root_path, undo_data)
    return True

def get_real_max_row(ws, col=1):
    """Find the actual last row with data in a given column."""
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, col).value is not None and str(ws.cell(r, col).value).strip() != "":
            return r
    return 0

def get_last_numeric_id(ws, col=1, start_row=2):
    """Get the last numeric ID in a column."""
    last_id = 0
    real_max = get_real_max_row(ws, col)
    for r in range(start_row, real_max + 1):
        val = ws.cell(r, col).value
        if val is not None:
            try:
                num = int(val)
                if num > last_id:
                    last_id = num
            except (ValueError, TypeError):
                pass
    return last_id

def copy_and_rename_file(src_path, dest_path, old_name, new_name, root_path, step_index):
    """Copies a file and does a raw string replacement of old_name to new_name in its contents."""
    if not os.path.exists(src_path):
        log(f"Error: Source file not found: {src_path}")
        return False

    try:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        
        with open(src_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        content = content.replace(old_name.lower(), new_name.lower())
        content = content.replace(old_name.capitalize(), new_name.capitalize())
        content = content.replace(old_name.upper(), new_name.upper())

        with open(dest_path, 'w', encoding='utf-8') as f:
            f.write(content)
            
        # Log for undo
        record_file_creation(root_path, step_index, dest_path)
            
        log(f"Successfully copied and patched {os.path.basename(dest_path)}")
        return True
    except Exception as e:
        log(f"Exception during file copy/patch: {e}")
        traceback.print_exc()
        return False

def inject_proto(proto_dir, src_evt, tgt_evt, root_path, step_index):
    log("Injecting Proto configs...")
    passed_proto = os.path.join(proto_dir, "passed_events.proto")
    storage_proto = os.path.join(proto_dir, "storagedata.proto")
    
    if not os.path.exists(passed_proto) or not os.path.exists(storage_proto):
        log("Proto files not found.")
        return False
        
    # Extract the full block belonging to the event (Bucket + any auxiliary messages)
    def extract_event_block(content, evt_name):
        msg_name = f"E{evt_name.capitalize()}Bucket"
        pattern = re.compile(r'message\s+' + msg_name + r'\s*\{')
        match = pattern.search(content)
        if not match: return None
        
        start_idx = match.start()
        
        # Look for the next event bucket to find the boundary of this event's code block
        # Matches newlines followed optionally by comments, then 'message E...Bucket'
        next_bucket_pattern = re.compile(r'\n\s*(?://.*?\n\s*)*message\s+E[A-Z]\w+Bucket\s*\{')
        next_match = next_bucket_pattern.search(content, match.end())
        
        if next_match:
            end_idx = next_match.start()
        else:
            end_idx = len(content)
            
        return content[start_idx:end_idx].strip()

    with open(passed_proto, 'r', encoding='utf-8') as f:
        passed_content = f.read()
        
    src_msg_name = f"E{src_evt.capitalize()}Bucket"
    tgt_msg_name = f"E{tgt_evt.capitalize()}Bucket"
    
    msg_block = extract_event_block(passed_content, src_evt)
    if not msg_block:
        log(f"Message {src_msg_name} not found in passed_events.proto. (Is it already in storagedata.proto?)")
        with open(storage_proto, 'r', encoding='utf-8') as f:
            msg_block = extract_event_block(f.read(), src_evt)
            if msg_block:
                log("Found source event in storagedata.proto instead. We will use that.")
            else:
                return False

    with open(storage_proto, 'r', encoding='utf-8') as f:
        storage_content = f.read()

    # If it already exists in target
    if extract_event_block(storage_content, tgt_evt):
        log(f"Target event message {tgt_msg_name} already exists in storagedata.proto. Skipping insertion.")
        return True

    if src_evt == tgt_evt:
        # Reopen Mode: Just append exactly as is
        new_msg_block = msg_block
    else:
        # Create New Event Mode: Rename inside the block
        new_msg_block = msg_block.replace(src_evt.lower(), tgt_evt.lower())
        new_msg_block = new_msg_block.replace(src_evt.capitalize(), tgt_evt.capitalize())
        new_msg_block = new_msg_block.replace(src_evt.upper(), tgt_evt.upper())
        
    # Create a backup before modifying
    shutil.copy2(storage_proto, storage_proto + ".bak")
    record_file_creation(root_path, step_index, storage_proto + ".bak") # Treat bak like a creation so it can be reverted
    
    # Append the block just before the end or at the end of the file
    storage_content += f"\n\n// Added by Event Automator: {tgt_evt}\n{new_msg_block}\n"
    
    with open(storage_proto, 'w', encoding='utf-8') as f:
        f.write(storage_content)
        
    # Also we should record the modification of storagedata.proto, but since we rely on the bak file for revert:
    log(f"Successfully injected proto message {tgt_msg_name} to storagedata.proto")
    return True

def inject_local_service_mgr(proj_root, tgt_evt, root_path, step_index):
    mgr_path = os.path.join(proj_root, r"develop\client\Skipbo\Assets\Lua\Game\Module\LocalService\LocalServiceMgr.lua")
    if not os.path.exists(mgr_path):
        log(f"Error: LocalServiceMgr.lua not found at {mgr_path}")
        return False
        
    with open(mgr_path, 'r', encoding='utf-8') as f:
        content = f.read()

    tgt_cap = tgt_evt.capitalize()
    require_str_base = f'local E{tgt_cap}LocalService = require "Game/Module/LocalService/Event/E{tgt_cap}LocalService"'
    require_str = f'{require_str_base}\n'
    
    bucket_str_base = f'E{tgt_cap}Bucket = E{tgt_cap}LocalService.New(LocalServiceMgr, "E{tgt_cap}Bucket"),'
    bucket_str = f'        {bucket_str_base}\n'

    # 1. Check if actively registered (line starts directly with 'local' not '--')
    if re.search(r'^\s*' + re.escape(require_str_base), content, re.MULTILINE):
        log(f"LocalServiceMgr: {tgt_evt} already registered and active. Skipping.")
        return True

    # 2. Check if commented out (line starts with '--')
    commented_req_pattern = r'^\s*--\s*' + re.escape(require_str_base)
    if re.search(commented_req_pattern, content, re.MULTILINE):
        log(f"Found commented out registration for {tgt_evt}. Uncommenting...")
        content = re.sub(commented_req_pattern, require_str_base, content, flags=re.MULTILINE)
        
        # Pattern to find the commented bucket line (e.g. `--        E...`)
        commented_bucket_pattern = r'^\s*--\s*(?:.*?)' + re.escape(bucket_str_base)
        content = re.sub(commented_bucket_pattern, r'        ' + bucket_str_base, content, flags=re.MULTILINE)
        
        shutil.copy2(mgr_path, mgr_path + ".bak")
        record_file_creation(root_path, step_index, mgr_path + ".bak")
        with open(mgr_path, 'w', encoding='utf-8') as f:
            f.write(content)
        log(f"Successfully uncommented {tgt_evt} in LocalServiceMgr.lua")
        return True

    # 3. If neither active nor commented, insert it as a brand new event

    # Insert require statement before PBReg
    if "local PBReg = " in content:
        content = content.replace('local PBReg = require "Game/Module/Protocol/PBReg"', require_str + 'local PBReg = require "Game/Module/Protocol/PBReg"')
    else:
        log("Error: Could not find 'local PBReg =' in LocalServiceMgr.lua")
        return False

    # Insert bucket inside ServiceDef
    find_str = "    }\n\n    LocalServiceMgr.LoginLocalService ="
    replace_str = bucket_str + find_str
    if find_str in content:
        content = content.replace(find_str, replace_str)
    else:
        log("Error: Could not find ServiceDef insertion point in LocalServiceMgr.lua")
        return False

    shutil.copy2(mgr_path, mgr_path + ".bak")
    record_file_creation(root_path, step_index, mgr_path + ".bak")
    
    with open(mgr_path, 'w', encoding='utf-8') as f:
        f.write(content)
        
    log(f"Successfully registered {tgt_evt} in LocalServiceMgr.lua")
    return True

def clone_workspace3_to_design(proj_root, step_index):
    import subprocess
    
    design_dir = os.path.join(proj_root, r"design\DesignData")
    workspace3 = os.path.join(design_dir, "WorkSpcae3")
    workspace_design = os.path.join(design_dir, "WorkSpcae_Design")
    
    if not os.path.exists(workspace3):
        log(f"Error: WorkSpcae3 not found at {workspace3}")
        return False
        
    try:
        # 1. SVN Update design directory
        log("Running SVN Update on DesignData...")
        subprocess.run(["svn", "cleanup", "."], cwd=design_dir, shell=True, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["svn", "up", ".", "--non-interactive", "--trust-server-cert"], cwd=design_dir, shell=True, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # 2. rmdir /s/q WorkSpcae_Design (same as original bat)
        if os.path.exists(workspace_design):
            log("Removing old WorkSpcae_Design directory...")
            subprocess.run(["cmd", "/c", "rmdir", "/s", "/q", workspace_design], check=False)
            
        # 3. xcopy WorkSpcae3 -> WorkSpcae_Design
        log("Cloning WorkSpcae3 -> WorkSpcae_Design...")
        shutil.copytree(workspace3, workspace_design)
        
        # Record creation for undo - FIX: Use provided root_path for central undo_log.json
        # NOTE: We need to pass the log_dir here, but it's not in the signature.
        # However, for Step 5 it might be called differently. Let's fix the call site or use a global.
        # For now, we'll assume the undo log is managed by the caller for this step,
        # or that the entire workspace_design can be removed if needed.
        # record_file_creation(design_dir, step_index, workspace_design) # Removed as per instruction
        
        # 4. SVN revert local utility bats in the new folder
        log("Reverting base bat files via SVN...")
        files_to_revert = [
            r"WorkSpcae_Design\PostConvert\Lua\lua_converter.lua",
            r"WorkSpcae_Design\copy_data.bat",
            r"WorkSpcae_Design\a_design_commit.bat",
            r"WorkSpcae_Design\a_design_commit_with_convert.bat"
        ]
        for f in files_to_revert:
            subprocess.run(["svn", "revert", f], cwd=design_dir, shell=True, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
        log("Successfully Cloned WorkSpace3 configuration to Design.")
        return True
    except Exception as e:
        log(f"Exception during Clone_2_Design: {e}")
        traceback.print_exc()
        return False

def clone_event_excel(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 6: Clone Event Excel configuration tables."""
    import openpyxl
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    backup_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\backup_xlsx")
    
    src_xlsx = src_evt.lower() + ".xlsx"
    tgt_xlsx = tgt_evt.lower() + ".xlsx"
    
    tgt_path = os.path.join(design_dir, tgt_xlsx)
    
    # Check if target already exists in design
    if os.path.exists(tgt_path):
        if src_evt == tgt_evt:
            log(f"Reopen: {tgt_xlsx} already exists in design directory. Skipping.")
            return True
        else:
            log(f"New Event: {tgt_xlsx} already exists in design directory. Skipping.")
            return True
    
    # Find the source file
    src_path = os.path.join(design_dir, src_xlsx)
    if not os.path.exists(src_path):
        # Try backup_xlsx
        src_path = os.path.join(backup_dir, src_xlsx)
        if not os.path.exists(src_path):
            log(f"Error: Source Excel {src_xlsx} not found in design or backup_xlsx directories.")
            return False
        log(f"Found source Excel in backup_xlsx: {src_xlsx}")
    else:
        log(f"Found source Excel in design: {src_xlsx}")
    
    try:
        if src_evt == tgt_evt:
            # ---- REOPEN MODE ----
            # Just copy the file from backup to design if it was missing
            if os.path.dirname(src_path) != design_dir:
                shutil.copy2(src_path, tgt_path)
                record_file_creation(root_path, step_index, tgt_path)
                log(f"Reopen: Copied {src_xlsx} from backup_xlsx to design.")
            else:
                log(f"Reopen: {src_xlsx} already in design directory.")
            return True
        else:
            # ---- NEW EVENT MODE ----
            # Copy the file with the new name
            shutil.copy2(src_path, tgt_path)
            record_file_creation(root_path, step_index, tgt_path)
            log(f"Copied {src_xlsx} -> {tgt_xlsx}")
            
            # Open with openpyxl and replace content
            log(f"Replacing event references inside {tgt_xlsx}...")
            # Use data_only=True to ensure formulas (especially for IDs) are converted to static values
            # This is critical for coder_convert to read the IDs correctly.
            wb = openpyxl.load_workbook(tgt_path, data_only=True)
            
            replacements = [
                (src_evt.lower(), tgt_evt.lower()),
                (src_evt.capitalize(), tgt_evt.capitalize()),
                (src_evt.upper(), tgt_evt.upper()),
            ]
            
            cells_changed = 0
            sheets_renamed = 0
            
            for ws in wb.worksheets:
                # Rename sheet if it contains the source event name
                for old, new in replacements:
                    if old in ws.title:
                        ws.title = ws.title.replace(old, new)
                        sheets_renamed += 1
                
                # Replace cell values
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original = cell.value
                            for old, new in replacements:
                                cell.value = cell.value.replace(old, new)
                            if cell.value != original:
                                cells_changed += 1
            
            wb.save(tgt_path)
            wb.close()
            
            log(f"Successfully cloned Excel: {tgt_xlsx} ({cells_changed} cells updated, {sheets_renamed} sheets renamed)")
            return True
    except Exception as e:
        log(f"Exception during Excel clone: {e}")
        traceback.print_exc()
        return False

def clone_event_descriptors(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 7: Copy event descriptor JSON(s) to app_client directory."""
    import glob
    
    app_client_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\descriptor\app_client")
    backup_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\descriptor\backup")
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    
    # 1. Find all source descriptor JSONs via glob
    src_pattern_app = os.path.join(app_client_dir, f"*{src_lower}*_convert.json")
    src_pattern_bak = os.path.join(backup_dir, f"*{src_lower}*_convert.json")
    
    src_files = glob.glob(src_pattern_app)
    source_location = "app_client"
    
    if not src_files:
        src_files = glob.glob(src_pattern_bak)
        source_location = "backup"
    
    if not src_files:
        log(f"Error: No descriptor JSONs matching '*{src_lower}*_convert.json' found in app_client or backup.")
        return False
    
    log(f"Found {len(src_files)} descriptor JSON(s) in {source_location}: {[os.path.basename(f) for f in src_files]}")
    
    try:
        for src_path in src_files:
            src_basename = os.path.basename(src_path)
            
            if src_evt == tgt_evt:
                # ---- REOPEN MODE ----
                tgt_basename = src_basename
                tgt_path = os.path.join(app_client_dir, tgt_basename)
                if os.path.exists(tgt_path):
                    log(f"Reopen: {tgt_basename} already in app_client. Skipping.")
                    continue
                shutil.copy2(src_path, tgt_path)
                record_file_creation(root_path, step_index, tgt_path)
                log(f"Reopen: Copied {src_basename} from {source_location} to app_client.")
            else:
                # ---- NEW EVENT MODE ----
                tgt_basename = src_basename.replace(src_lower, tgt_lower)
                tgt_path = os.path.join(app_client_dir, tgt_basename)
                
                if os.path.exists(tgt_path):
                    log(f"New Event: {tgt_basename} already exists. Skipping.")
                    continue
                
                # Read source, replace content, write target
                with open(src_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                content = content.replace(src_evt.lower(), tgt_evt.lower())
                content = content.replace(src_evt.capitalize(), tgt_evt.capitalize())
                content = content.replace(src_evt.upper(), tgt_evt.upper())
                
                with open(tgt_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                record_file_creation(root_path, step_index, tgt_path)
                log(f"Created {tgt_basename} with content renamed from {src_basename}")
        
        log("Successfully cloned all event descriptor JSON(s).")
        return True
    except Exception as e:
        log(f"Exception during descriptor clone: {e}")
        traceback.print_exc()
        return False

def inject_convert_references(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 8: Modify convert_layout.lua and convert.json to register the new event."""
    import json
    
    design_base = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design")
    convert_json_path = os.path.join(design_base, r"descriptor\app_client\convert.json")
    convert_layout_path = os.path.join(design_base, r"PostConvert\Lua\convert_layout.lua")
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    
    success_json = False
    success_lua = False
    
    # ========== Part 1: convert.json ==========
    try:
        if not os.path.exists(convert_json_path):
            log(f"Error: convert.json not found at {convert_json_path}")
            return False
        
        with open(convert_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Find the const_event.json Director entry
        const_event_dir = None
        for director in data.get("Directors", []):
            if director.get("FileName") == "const_event.json":
                const_event_dir = director
                break
        
        if not const_event_dir:
            log("Warning: 'const_event.json' Director not found in convert.json. Skipping JSON injection.")
        else:
            sheets = const_event_dir.get("ExcelSheets", [])
            tgt_entry = {"ExcelName": tgt_lower, "SheetName": f"{tgt_lower}_const"}
            
            # Check if already exists
            already_exists = any(
                s.get("ExcelName") == tgt_lower for s in sheets
            )
            
            if already_exists:
                log(f"convert.json: {tgt_lower} already registered in const_event.json. Skipping.")
            else:
                # Backup
                backup_path = convert_json_path + ".bak"
                shutil.copy2(convert_json_path, backup_path)
                record_file_creation(root_path, step_index, backup_path)
                
                sheets.append(tgt_entry)
                
                with open(convert_json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                
                log(f"convert.json: Added {tgt_lower} to const_event.json ExcelSheets.")
            
            success_json = True
    except Exception as e:
        log(f"Exception updating convert.json: {e}")
        traceback.print_exc()
        return False
    
    # ========== Part 2: convert_layout.lua ==========
    try:
        import glob as _glob
        import re
        
        if not os.path.exists(convert_layout_path):
            log(f"Error: convert_layout.lua not found at {convert_layout_path}")
            return False
        
        with open(convert_layout_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Check if target event already has active (uncommented) entries
        has_active = bool(re.search(rf'(?m)^\s*\{{.*json\s*=\s*"{tgt_lower}_.*"', content))
        has_commented = bool(re.search(rf'(?m)^\s*--.*{tgt_lower}_', content))
        
        if has_active:
            log(f"convert_layout.lua: {tgt_lower} entries already exist. Skipping.")
            success_lua = True
        elif has_commented:
            # Uncomment the lines
            lines = content.split('\n')
            count = 0
            for i, line in enumerate(lines):
                stripped = line.strip()
                if tgt_lower in stripped and stripped.startswith('--') and 'json' in stripped:
                    new_stripped = stripped.lstrip('-').lstrip()
                    lines[i] = line.replace(stripped, new_stripped)
                    count += 1
            
            content = '\n'.join(lines)
            backup_path = convert_layout_path + ".bak"
            shutil.copy2(convert_layout_path, backup_path)
            record_file_creation(root_path, step_index, backup_path)
            
            with open(convert_layout_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            log(f"convert_layout.lua: Uncommented {count} entries for {tgt_lower}.")
            success_lua = True
        else:
            # Need to generate entries — read the target's _convert.json
            app_client_dir = os.path.join(design_base, r"descriptor\app_client")
            backup_desc_dir = os.path.join(design_base, r"descriptor\backup")
            
            # Find the target _convert.json file(s)
            tgt_convert_files = _glob.glob(os.path.join(app_client_dir, f"*{tgt_lower}*_convert.json"))
            if not tgt_convert_files:
                tgt_convert_files = _glob.glob(os.path.join(backup_desc_dir, f"*{tgt_lower}*_convert.json"))
            
            if not tgt_convert_files:
                log(f"Warning: No _convert.json found for {tgt_lower}. Cannot generate convert_layout entries.")
                success_lua = True
            else:
                # Collect all FileNames from the target's _convert.json(s)
                tgt_filenames = []
                for cf in tgt_convert_files:
                    with open(cf, 'r', encoding='utf-8') as f:
                        cdata = json.load(f)
                    for director in cdata.get("Directors", []):
                        fn = director.get("FileName", "")
                        if fn:
                            tgt_filenames.append(fn)
                
                log(f"Read {len(tgt_filenames)} FileNames from target _convert.json(s).")
                
                # Build a lookup: for each existing entry in convert_layout.lua,
                # map the suffix (FileName minus event prefix) to (section, merge_as_suffix)
                # We parse the entire lua to understand sections and their entries
                lines = content.split('\n')
                
                # Identify which section each line belongs to
                current_section = None
                section_end_lines = {}  # section -> last line index with an entry
                suffix_to_info = {}     # suffix -> (section, merge_as_suffix, sample_line)
                
                for i, line in enumerate(lines):
                    stripped = line.strip()
                    # Detect section headers like "default_map = {"
                    section_match = re.match(r'(\w+)\s*=\s*\{', stripped)
                    if section_match:
                        current_section = section_match.group(1)
                    
                    # Detect json entries
                    entry_match = re.match(r'.*json\s*=\s*"([^"]+)".*merge_as\s*=\s*"([^"]+)"', stripped)
                    if entry_match and current_section and not stripped.startswith('--'):
                        json_name = entry_match.group(1)  # e.g., "blocks2601_scene.json"
                        merge_as = entry_match.group(2)   # e.g., "blocks2601.scene"
                        
                        # Track section end lines
                        section_end_lines[current_section] = i
                        
                        # Extract the suffix: strip any known event prefix
                        # Try to find the event name by matching against merge_as prefix
                        dot_idx = merge_as.find('.')
                        if dot_idx > 0:
                            evt_prefix = merge_as[:dot_idx]  # e.g., "blocks2601"
                            merge_suffix = merge_as[dot_idx+1:]  # e.g., "scene"
                            
                            # The json filename suffix: strip event prefix + underscore + .json
                            if json_name.startswith(evt_prefix + '_'):
                                file_suffix = json_name[len(evt_prefix)+1:]  # e.g., "scene.json"
                            else:
                                file_suffix = json_name
                            
                            # Store: file_suffix -> (section, merge_suffix, line_template)
                            if file_suffix not in suffix_to_info:
                                suffix_to_info[file_suffix] = (current_section, merge_suffix, line)
                
                log(f"Built suffix lookup with {len(suffix_to_info)} known patterns.")
                
                # Now generate entries for each target FileName
                # Group by section for insertion
                section_entries = {}  # section -> list of new lines
                
                for fn in tgt_filenames:
                    # Derive the file suffix: strip target event prefix
                    if fn.startswith(tgt_lower + '_'):
                        file_suffix = fn[len(tgt_lower)+1:]  # e.g., "scene.json"
                    else:
                        file_suffix = fn
                    
                    # Look up the suffix in our mapping
                    if file_suffix in suffix_to_info:
                        section, merge_suffix, sample_line = suffix_to_info[file_suffix]
                    else:
                        # Fallback: use default_map, derive merge_suffix from file_suffix
                        section = "default_map"
                        merge_suffix = file_suffix.replace('.json', '')
                    
                    merge_as = f"{tgt_lower}.{merge_suffix}"
                    new_line = f'        {{json = "{fn}", merge_as = "{merge_as}"}},'
                    
                    if section not in section_entries:
                        section_entries[section] = []
                    section_entries[section].append(new_line)
                
                # Insert entries into convert_layout.lua, at the end of each section
                lines = content.split('\n')
                offset = 0
                
                backup_path = convert_layout_path + ".bak"
                shutil.copy2(convert_layout_path, backup_path)
                record_file_creation(root_path, step_index, backup_path)
                
                total_injected = 0
                for section, entries in section_entries.items():
                    if section in section_end_lines:
                        insert_idx = section_end_lines[section] + 1 + offset
                        new_block = [''] + entries
                        for j, entry in enumerate(new_block):
                            lines.insert(insert_idx + j, entry)
                        offset += len(new_block)
                        total_injected += len(entries)
                        log(f"  Injected {len(entries)} entries into '{section}' section.")
                    else:
                        # Section not found, inject into default_map as fallback
                        if "default_map" in section_end_lines:
                            insert_idx = section_end_lines["default_map"] + 1 + offset
                            new_block = [''] + entries
                            for j, entry in enumerate(new_block):
                                lines.insert(insert_idx + j, entry)
                            offset += len(new_block)
                            total_injected += len(entries)
                            log(f"  Injected {len(entries)} entries into 'default_map' (fallback for '{section}').")
                
                content = '\n'.join(lines)
                
                with open(convert_layout_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                log(f"convert_layout.lua: Injected {total_injected} total entries for {tgt_lower}.")
                success_lua = True
    except Exception as e:
        log(f"Exception updating convert_layout.lua: {e}")
        traceback.print_exc()
        return False
    
    return success_json and success_lua

def update_bi_excel(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 9: Update bi.xlsx with target event tracking entries."""
    import openpyxl
    import re
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    bi_path = os.path.join(design_dir, "bi.xlsx")
    
    if not os.path.exists(bi_path):
        log(f"Error: bi.xlsx not found at {bi_path}")
        return False
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()
    # PascalCase: first letter upper, rest as-is from the original
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    
    def swap_event_name(value):
        """Replace all case variants of the source event name with the target."""
        if value is None:
            return None
        s = str(value)
        s = s.replace(src_upper, tgt_upper)
        s = s.replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal)
        # Also try the original casing
        s = s.replace(src_evt, tgt_evt)
        return s
    
    try:
        # Backup first
        backup_path = bi_path + ".bak"
        if os.path.exists(backup_path):
            import stat
            os.chmod(backup_path, stat.S_IWRITE)
            os.remove(backup_path)
        shutil.copy2(bi_path, backup_path)
        record_file_creation(root_path, step_index, backup_path)
        
        wb = openpyxl.load_workbook(bi_path)
        modified = False
        
        # ========== Sheet 1: bi_reason ==========
        if "bi_reason" in wb.sheetnames:
            ws = wb["bi_reason"]
            # Check if target already exists
            already_exists = False
            for r in range(1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val and tgt_lower in str(val).lower():
                    already_exists = True
                    break
            
            if already_exists:
                log(f"bi_reason: {tgt_evt} already exists. Skipping.")
            else:
                # Find source event row
                src_row = None
                for r in range(1, ws.max_row + 1):
                    val = ws.cell(r, 1).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = ws.max_row + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_event_name(old_val))
                    log(f"bi_reason: Added row {new_row} for {tgt_evt}.")
                    modified = True
                else:
                    log(f"bi_reason: Source event {src_evt} not found. Skipping.")
        
        # ========== Sheet 2: 广告打点常量表 ==========
        ad_sheet_name = "广告打点常量表"
        if ad_sheet_name in wb.sheetnames:
            ws = wb[ad_sheet_name]
            # Check if target already exists
            already_exists = False
            for r in range(1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val and tgt_upper in str(val).upper():
                    already_exists = True
                    break
            
            if already_exists:
                log(f"{ad_sheet_name}: {tgt_evt} already exists. Skipping.")
            else:
                # Collect source rows grouped by prefix (SITE vs ENTRANCE)
                groups = {}  # prefix -> [(row_idx, [cell_values])]
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(r, 1).value
                    if val and src_upper in str(val).upper():
                        enum_val = str(val)
                        # Determine group: AD_BI_SITE_ or AD_BI_ENTRANCE_
                        if 'AD_BI_SITE_' in enum_val:
                            group_key = 'AD_BI_SITE'
                        elif 'AD_BI_ENTRANCE_' in enum_val:
                            group_key = 'AD_BI_ENTRANCE'
                        else:
                            group_key = 'OTHER'
                        
                        row_data = []
                        for c in range(1, ws.max_column + 1):
                            row_data.append(ws.cell(r, c).value)
                        
                        if group_key not in groups:
                            groups[group_key] = []
                        groups[group_key].append((r, row_data))
                
                # For each group, find the last row of that group and insert after
                for group_key, rows_data in groups.items():
                    # Find the last row in the sheet that belongs to this group
                    last_group_row = 0
                    for r in range(2, ws.max_row + 1):
                        val = ws.cell(r, 1).value
                        if val and group_key in str(val):
                            last_group_row = r
                    
                    if last_group_row == 0:
                        last_group_row = ws.max_row
                    
                    # Insert new rows after the last group row
                    # We need to use insert_rows to shift existing data down
                    insert_at = last_group_row + 1
                    num_new = len(rows_data)
                    ws.insert_rows(insert_at, num_new)
                    
                    for i, (_, cell_values) in enumerate(rows_data):
                        for c in range(1, len(cell_values) + 1):
                            ws.cell(insert_at + i, c, swap_event_name(cell_values[c-1]))
                    
                    log(f"{ad_sheet_name}: Inserted {num_new} rows for {tgt_evt} in {group_key} group at row {insert_at}.")
                    modified = True
        
        # ========== Sheet 3: bi_event_names ==========
        if "bi_event_names" in wb.sheetnames:
            ws = wb["bi_event_names"]
            # Check if target already exists
            already_exists = False
            for r in range(1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val and tgt_lower in str(val).lower():
                    already_exists = True
                    break
            
            if already_exists:
                log(f"bi_event_names: {tgt_evt} already exists. Skipping.")
            else:
                # Find source event row
                src_row = None
                for r in range(1, ws.max_row + 1):
                    val = ws.cell(r, 1).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = ws.max_row + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_event_name(old_val))
                    log(f"bi_event_names: Added row {new_row} for {tgt_evt}.")
                    modified = True
                else:
                    log(f"bi_event_names: Source event {src_evt} not found. Skipping.")
        
        if modified:
            import stat
            os.chmod(bi_path, stat.S_IWRITE | stat.S_IREAD)
            wb.save(bi_path)
            log(f"bi.xlsx saved successfully.")
        else:
            log(f"bi.xlsx: No changes were needed.")
        
        return True
    except Exception as e:
        log(f"Exception updating bi.xlsx: {e}")
        traceback.print_exc()
        return False

def load_extra_args(root_path):
    """Load extra arguments from extra_args.json written by the C# GUI."""
    args_path = os.path.join(root_path, "extra_args.json")
    if os.path.exists(args_path):
        with open(args_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def update_events_excel(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 10: Update events.xlsx with target event registration data."""
    import openpyxl
    import stat
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    events_path = os.path.join(design_dir, "events.xlsx")
    
    if not os.path.exists(events_path):
        log(f"Error: events.xlsx not found at {events_path}")
        return False
    
    extra = load_extra_args(root_path)
    event_id = extra.get("event_id", "")
    start_time = extra.get("start_time", "")
    end_time = extra.get("end_time", "")
    near_end_time = extra.get("near_end_time", "")
    close_time = extra.get("close_time", "")
    is_reopen = extra.get("is_reopen", "0") == "1"
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    
    def swap_name(value):
        if value is None:
            return None
        s = str(value)
        # Standard string swaps
        if src_evt:
            s = s.replace(src_upper, tgt_upper)
            s = s.replace(src_lower, tgt_lower)
            s = s.replace(src_pascal, tgt_pascal)
            s = s.replace(src_evt, tgt_evt)
        # Fallback for ID-like columns (e.g. 20250701 -> 20260501)
        src_digits = "".join(filter(str.isdigit, src_evt))
        tgt_digits = "".join(filter(str.isdigit, tgt_evt))
        if len(src_digits) >= 4 and len(tgt_digits) >= 4:
            s = s.replace("20" + src_digits[:4], "20" + tgt_digits[:4])
        return s
    
    def find_src_event_id(ws):
        """Find the Event ID used by the source event in a sheet."""
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if val and src_lower in str(val).lower():
                    # Return col 1 value of this row as the event ID
                    return str(ws.cell(r, 1).value) if ws.cell(r, 1).value else None
        return None
    
    def target_exists(ws):
        """Check if target event already has entries in a sheet."""
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if val and tgt_lower in str(val).lower():
                    return True
        return False


    
    try:
        # Backup
        backup_path = events_path + ".bak"
        if os.path.exists(backup_path):
            os.chmod(backup_path, stat.S_IWRITE)
            os.remove(backup_path)
        shutil.copy2(events_path, backup_path)
        record_file_creation(root_path, step_index, backup_path)
        
        wb = openpyxl.load_workbook(events_path)
        modified = False
        src_event_id = None  # Will be discovered from data
        
        # ========== Sheet 1: events (main) ==========
        if "events" in wb.sheetnames:
            ws = wb["events"]
            if target_exists(ws):
                log("events: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(r, c).value
                        if val and src_lower in str(val).lower():
                            src_row = r
                            break
                    if src_row:
                        break
                
                if src_row:
                    src_event_id = str(ws.cell(src_row, 1).value)
                    last_row_idx = real_max
                    new_row = last_row_idx + 1
                    
                    # Set the last row's control switch to FALSE (col 6) - as it's the "previous" event
                    if last_row_idx >= 2:
                        ws.cell(last_row_idx, 6, "FALSE")
                        log(f"events: Set previous event (row {last_row_idx}) control switch to FALSE.")
                    
                    # Copy all columns from template source
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    # Override specific columns
                    # Col 1: Event ID from GUI input
                    if event_id:
                        ws.cell(new_row, 1, int(event_id))
                    
                    # Col 4: BI Title = first 6 digits of ID + target name
                    if event_id and len(event_id) >= 6:
                        bi_title = event_id[:6] + tgt_pascal
                        ws.cell(new_row, 4, bi_title)
                    
                    # Col 5: Localization key swap
                    ws.cell(new_row, 5, f"{tgt_upper}_EVENTNAME")
                    
                    # Col 6: Control switch = TRUE for new event
                    ws.cell(new_row, 6, "TRUE")
                    
                    # Col 8-11: Dates from GUI
                    if start_time:
                        ws.cell(new_row, 8, start_time)
                    if end_time:
                        ws.cell(new_row, 9, end_time)
                    if near_end_time:
                        ws.cell(new_row, 10, near_end_time)
                    if close_time:
                        ws.cell(new_row, 11, close_time)
                    
                    # Col 23: Previous event ref = alias of the LAST row (the one we just set to FALSE)
                    if last_row_idx >= 2:
                        last_alias = ws.cell(last_row_idx, 2).value
                        if last_alias:
                            ws.cell(new_row, 23, str(last_alias))
                    
                    # Col 27: Drop item ID - leave source value for now, updated in Step 12
                    
                    log(f"events: Added row {new_row} for {tgt_evt}.")
                    modified = True
                else:
                    log(f"events: Source event {src_evt} not found.")
        
        # Discover src_event_id if not yet found
        if not src_event_id and "events" in wb.sheetnames:
            src_event_id = find_src_event_id(wb["events"])
        
        # ========== Sheet 2: event_type ==========
        if "event_type" in wb.sheetnames:
            ws = wb["event_type"]
            if target_exists(ws):
                log("event_type: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 3).value  # Col 3 is the enum
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = real_max + 1
                    last_id = get_last_numeric_id(ws, 1)
                    
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    # Col 1: Auto-increment type ID
                    ws.cell(new_row, 1, last_id + 1)
                    
                    log(f"event_type: Added row {new_row} (ID={last_id+1}) for {tgt_evt}.")
                    modified = True
        
        # ========== Sheet 3: event_entrance ==========
        if "event_entrance" in wb.sheetnames:
            ws = wb["event_entrance"]
            if target_exists(ws):
                log("event_entrance: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 2).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = real_max + 1
                    last_id = get_last_numeric_id(ws, 1)
                    
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    ws.cell(new_row, 1, last_id + 1)
                    
                    log(f"event_entrance: Added row {new_row} (ID={last_id+1}) for {tgt_evt}.")
                    modified = True
        
        # ========== Sheet 4: event_score ==========
        if "event_score" in wb.sheetnames:
            ws = wb["event_score"]
            if target_exists(ws):
                log("event_score: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 2).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = real_max + 1
                    last_id = get_last_numeric_id(ws, 1)
                    
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    ws.cell(new_row, 1, last_id + 1)
                    
                    log(f"event_score: Added row {new_row} (ID={last_id+1}) for {tgt_evt}.")
                    modified = True
        
        # ========== Sheet 5: event_pre ==========
        if "event_pre" in wb.sheetnames:
            ws = wb["event_pre"]
            if target_exists(ws):
                log("event_pre: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 2).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = real_max + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    # Col 1: Use new event ID
                    if event_id:
                        ws.cell(new_row, 1, int(event_id))
                    
                    log(f"event_pre: Added row {new_row} for {tgt_evt}.")
                    modified = True
        
        # ========== Sheet 6: event_post ==========
        if "event_post" in wb.sheetnames:
            ws = wb["event_post"]
            if target_exists(ws):
                log("event_post: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 2).value
                    if val and src_lower in str(val).lower():
                        src_row = r
                        break
                
                if src_row:
                    new_row = real_max + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    if event_id:
                        ws.cell(new_row, 1, int(event_id))
                    
                    log(f"event_post: Added row {new_row} for {tgt_evt}.")
                    
                    # Update switches: last two = TRUE, rest = FALSE
                    # new_row is the last row.
                    for r in range(2, new_row + 1):
                        # Switch is in Column 3
                        # We want row new_row and new_row-1 to be TRUE
                        if r >= (new_row - 1):
                            ws.cell(r, 3, "TRUE")
                        else:
                            ws.cell(r, 3, "FALSE")
                    log(f"event_post: Set TRUE for last two rows in switch column.")
                    
                    modified = True
        
        # ========== Sheet 7: \u6d3b\u52a8\u5176\u4ed6 ==========
        other_sheet = "\u6d3b\u52a8\u5176\u4ed6"
        if other_sheet in wb.sheetnames:
            ws = wb[other_sheet]
            if target_exists(ws):
                log("activity_other: Target already exists. Skipping.")
            else:
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(r, c).value
                        if val and src_lower in str(val).lower():
                            src_row = r
                            break
                    if src_row:
                        break
                
                if src_row:
                    new_row = real_max + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    if event_id:
                        ws.cell(new_row, 1, int(event_id))
                    
                    log(f"activity_other: Added row {new_row} for {tgt_evt}.")
                    
                    # New adjustment: Comment out all rows except the last two
                    # After adding, new_row is the last row.
                    # We need rows (new_row) and (new_row-1) uncommented.
                    # Everything from 2 up to (new_row-2) should be commented.
                    for r in range(2, new_row - 1):
                        val = ws.cell(r, 1).value
                        if val is not None:
                            s = str(val)
                            if not s.startswith("//"):
                                ws.cell(r, 1, "//" + s)
                                log(f"activity_other: Commented out old row {r}.")
                    
                    modified = True
        
        # ========== Sheet 8: \u6d3b\u52a8\u6210\u5c31 ==========
        achieve_sheet = "\u6d3b\u52a8\u6210\u5c31"
        if achieve_sheet in wb.sheetnames:
            ws = wb[achieve_sheet]
            if target_exists(ws):
                log("activity_achievement: Target already exists. Skipping.")
            else:
                # For achievements, we need a reference row - try source first, then any recent event
                src_row = None
                real_max = get_real_max_row(ws)
                for r in range(2, real_max + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(r, c).value
                        if val and src_lower in str(val).lower():
                            src_row = r
                            break
                    if src_row:
                        break
                
                # If source not found, use the last row as template
                if not src_row and real_max >= 2:
                    src_row = real_max
                    log(f"activity_achievement: Source not found, using last row {src_row} as template.")
                
                if src_row:
                    new_row = real_max + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(src_row, c).value
                        ws.cell(new_row, c, swap_name(old_val))
                    
                    # Col 1: Event ID
                    if event_id:
                        ws.cell(new_row, 1, int(event_id))
                    
                    # Col 2: Achievement type derived from year
                    # If ID is 20260501 -> year is 2026 -> value 26
                    # If ID is 20270501 -> year is 2027 -> value 27
                    if event_id and len(event_id) >= 4:
                        year_suffix = int(event_id[:4]) - 2000  # 2026 -> 26
                        ws.cell(new_row, 2, year_suffix)
                    
                    # Col 3: Switch = TRUE
                    ws.cell(new_row, 3, "TRUE")
                    
                    # Col 4: Icon - swap name
                    ws.cell(new_row, 4, f"{tgt_pascal}活动成就" if not ws.cell(src_row, 4).value else swap_name(ws.cell(src_row, 4).value))
                    
                    # Col 5: Localization key
                    ws.cell(new_row, 5, f"ACHIEVEMENT_{tgt_upper}")
                    
                    # Col 6: DeepLink - auto increment
                    last_deeplink = get_last_numeric_id(ws, 6)
                    ws.cell(new_row, 6, last_deeplink + 1)
                    
                    log(f"activity_achievement: Added row {new_row} for {tgt_evt}.")
                    modified = True
        
        # ========== Sheet 9: \u6d3b\u52a8buff\u5e38\u91cf ==========
        buff_sheet = "\u6d3b\u52a8buff\u5e38\u91cf"
        buff_backup_sheet = "buff\u8001\u6d3b\u52a8\u5907\u4efd"
        if buff_sheet in wb.sheetnames and buff_backup_sheet in wb.sheetnames:
            ws = wb[buff_sheet]
            ws_backup = wb[buff_backup_sheet]
            
            # Step 1: Collect all event groups currently in the main sheet
            # A group consists of a header "//\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8(Name)" and rows starting with 202xxxxx
            main_groups = []
            current_group = None
            
            for r in range(1, ws.max_row + 1):
                c1 = ws.cell(r, 1).value
                if not c1:
                    if current_group:
                        main_groups.append(current_group)
                        current_group = None
                    continue
                
                s1 = str(c1)
                # Header row: //\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8(Name)
                if s1.startswith("//\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8"):
                    if current_group:
                        main_groups.append(current_group)
                    
                    name_match = re.search(r'\(([^)]+)\)', s1)
                    evt_name = name_match.group(1) if name_match else "Unknown"
                    current_group = {
                        'name': evt_name,
                        'rows': [[ws.cell(r, cc).value for cc in range(1, ws.max_column + 1)]]
                    }
                elif current_group and s1.startswith("202"):
                    # Data row belonging to the event
                    current_group['rows'].append([ws.cell(r, cc).value for cc in range(1, ws.max_column + 1)])
                else:
                    # Some other row: beginner or something else
                    if current_group:
                        main_groups.append(current_group)
                        current_group = None
            
            if current_group:
                main_groups.append(current_group)
            
            # Step 2: Move missing event groups from main to backup
            # Collect existing event names in backup
            backup_event_names = set()
            for r in range(1, ws_backup.max_row + 1):
                c1 = ws_backup.cell(r, 1).value
                if c1 and str(c1).startswith("//\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8"):
                    name_match = re.search(r'\(([^)]+)\)', str(c1))
                    if name_match:
                        backup_event_names.add(name_match.group(1).lower())
            
            for group in main_groups:
                if group['name'].lower() not in backup_event_names:
                    log(f"buff_constants: Backing up event {group['name']} to {buff_backup_sheet}")
                    for row_vals in group['rows']:
                        next_bk_row = get_real_max_row(ws_backup) + 1
                        for c_idx, val in enumerate(row_vals):
                            ws_backup.cell(next_bk_row, c_idx + 1, val)
                    modified = True
            
            # Step 3: Clear Main sheet of all bi-weekly events before injecting target
            # Skip clearing/injection if target already in main sheet unless we want a rebuild
            if target_exists(ws):
                log("buff_constants: Target already exists in main sheet. Skipping injection.")
            else:
                rows_to_delete = []
                for r in range(1, ws.max_row + 1):
                    c1 = ws.cell(r, 1).value
                    if c1:
                        s1 = str(c1)
                        if s1.startswith("//\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8") or s1.startswith("202"):
                            rows_to_delete.append(r)
                
                # Deleting from bottom up to preserve indices
                for r in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(r)
                log(f"buff_constants: Removed {len(rows_to_delete)} bi-weekly rows from main sheet.")
                
                # Step 4: Inject target event from backup
                src_template_rows = []
                src_id_prefix = None
                
                # Search backup for source event
                found_src = False
                for r in range(1, ws_backup.max_row + 1):
                    c1 = ws_backup.cell(r, 1).value
                    if c1 and str(c1).startswith("//\u6b63\u5e38\u53cc\u5468\u6d3b\u52a8") and src_lower in str(c1).lower():
                        # Found it!
                        found_src = True
                        src_template_rows.append([ws_backup.cell(r, cc).value for cc in range(1, ws_backup.max_column + 1)])
                        # Collect following 202xxxxx rows
                        nr = r + 1
                        while nr <= ws_backup.max_row:
                            nc1 = ws_backup.cell(nr, 1).value
                            if nc1 and str(nc1).startswith("202"):
                                data_row = [ws_backup.cell(nr, cc).value for cc in range(1, ws_backup.max_column + 1)]
                                if not src_id_prefix:
                                    src_id_prefix = str(nc1).split('_')[0]
                                src_template_rows.append(data_row)
                                nr += 1
                            else:
                                break
                        break
                
                if found_src and src_template_rows:
                    log(f"buff_constants: Templating {tgt_evt} from {src_evt} found in backup.")
                    
                    # Instead of append, use explicit row to avoid gaps from delete_rows residual max_row
                    curr_row = get_real_max_row(ws) + 1
                    
                    for row_data in src_template_rows:
                        for c_idx, val in enumerate(row_data):
                            new_val = swap_name(val)
                            # Replacement for Col 1 (Program Enum with ID prefix)
                            if c_idx == 0 and new_val and src_id_prefix and event_id:
                                if str(new_val).startswith(src_id_prefix):
                                    new_val = event_id + str(new_val)[len(src_id_prefix):]
                            ws.cell(curr_row, c_idx + 1, new_val)
                        curr_row += 1
                    modified = True
                else:
                    log(f"buff_constants: Source event {src_evt} not found in backup sheet.")
        
        if modified:
            os.chmod(events_path, stat.S_IWRITE | stat.S_IREAD)
            wb.save(events_path)
            log(f"events.xlsx saved successfully.")
        else:
            log(f"events.xlsx: No changes were needed.")
        
        return True
    except Exception as e:
        log(f"Exception updating events.xlsx: {e}")
        traceback.print_exc()
        return False

def update_item_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """
    Step 12: Update item.xlsx (Item definitions)
    """
    log(f"Step 12: Updating item.xlsx for {tgt_evt} (cloning from {src_evt})")
    
    design_data_path = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_data_path, "item.xlsx")
    
    if not os.path.exists(file_path):
        log(f"item.xlsx not found at {file_path}")
        return False
        
    try:
        import openpyxl
        # Backup
        bak_path = file_path + ".bak"
        if not os.path.exists(bak_path):
            shutil.copy2(file_path, bak_path)
            record_file_creation(log_dir, step_index, bak_path)
        
        # Make writable
        import stat
        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        # 1. Prepare name swap strings
        src_upper = src_evt.upper()
        tgt_upper = tgt_evt.upper()
        src_lower = src_evt.lower()
        tgt_lower = tgt_evt.lower()
        src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else ""
        tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else ""
        
        # Extract digits from src/tgt names (e.g. 2507, 2605)
        src_digits = "".join(filter(str.isdigit, src_evt))
        tgt_digits = "".join(filter(str.isdigit, tgt_evt))
        
        def swap_name(value):
            if value is None: return None
            s = str(value)
            if src_evt:
                s = s.replace(src_upper, tgt_upper)
                s = s.replace(src_lower, tgt_lower)
                s = s.replace(src_pascal, tgt_pascal)
                s = s.replace(src_evt, tgt_evt)
            if src_digits and tgt_digits:
                s = s.replace(src_digits, tgt_digits)
            return s

        # ========== Sheet: item ==========
        if "item" in wb.sheetnames:
            ws = wb["item"]
            real_max = get_real_max_row(ws)
            
            # Find source items
            src_rows = []
            for r in range(2, real_max + 1):
                alias_val = str(ws.cell(r, 2).value or "").lower()
                name_val = str(ws.cell(r, 5).value or "").lower()
                
                match = False
                if src_evt.lower() in alias_val or src_evt.lower() in name_val:
                    match = True
                
                if match:
                    src_rows.append(r)
            
            if src_rows:
                log(f"item_sheet: Found {len(src_rows)} items for {src_evt}. Cloning...")
                
                # Group source items by their ID range and find current max ID in the sheet for those ranges
                max_ids_by_range = {}
                for r in range(2, real_max + 1):
                    val = ws.cell(r, 1).value
                    if val:
                        s_val = str(val).strip()
                        if s_val.startswith("//"): s_val = s_val[2:].strip()
                        if s_val.isdigit():
                            num = int(s_val)
                            range_key = num // 1000
                            if range_key not in max_ids_by_range or num > max_ids_by_range[range_key]:
                                max_ids_by_range[range_key] = num
                
                id_map = {}
                new_rows_data = []
                for src_r in src_rows:
                    row_data = []
                    for c in range(1, ws.max_column + 1):
                        row_data.append(ws.cell(src_r, c).value)
                    
                    old_id_str = str(row_data[0]).strip()
                    if old_id_str.startswith("//"): old_id_str = old_id_str[2:].strip()
                    
                    new_id = -1
                    if old_id_str.isdigit():
                        old_id = int(old_id_str)
                        range_key = old_id // 1000
                        new_id = max_ids_by_range.get(range_key, old_id) + 1
                        max_ids_by_range[range_key] = new_id
                        row_data[0] = new_id
                        id_map[old_id] = new_id
                        log(f"item_sheet: Map ID {old_id} -> {new_id}")
                    else:
                        row_data[0] = swap_name(row_data[0])
                        new_id = 99999999 # Fallback
                    
                    for i in range(1, len(row_data)):
                        row_data[i] = swap_name(row_data[i])
                        
                    new_rows_data.append((new_id, row_data))
                
                new_rows_data.sort(key=lambda x: x[0], reverse=True)
                for new_id, row_data in new_rows_data:
                    best_row = 1
                    max_id_seen = -1
                    for r in range(2, ws.max_row + 1):
                        val = ws.cell(r, 1).value
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str.startswith("//"): val_str = val_str[2:].strip()
                            if val_str.isdigit():
                                current_id = int(val_str)
                                if current_id < new_id and current_id > max_id_seen:
                                    max_id_seen = current_id
                                    best_row = r
                    
                    insert_pos = best_row + 1
                    ws.insert_rows(insert_pos)
                    for c_idx, val in enumerate(row_data):
                        ws.cell(insert_pos, c_idx + 1, val)
                    log(f"item_sheet: Inserted item {new_id} at row {insert_pos}")
                    
                modified = True
            else:
                log(f"item_sheet: No source items found for {src_evt}.")
                id_map = {}
        
        if modified:
            wb.save(file_path)
            log(f"item.xlsx updated for {tgt_evt}.")
            return True, id_map
        else:
            log(f"item.xlsx: No changes were needed.")
            return True, {}
            
    except Exception as e:
        log(f"Error updating item.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False, {}


def sync_events_excel_item_ids(proj_root, tgt_evt, id_map, log_dir, step_index):
    """Update events.xlsx with the new '掉落的道具ID' from item.xlsx sync."""
    import openpyxl
    import os
    import stat
    
    events_path = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design\events.xlsx")
    if not os.path.exists(events_path):
        log(f"Warning: events.xlsx not found for ID sync at {events_path}")
        return True
        
    try:
        os.chmod(events_path, stat.S_IWRITE | stat.S_IREAD)
        wb = openpyxl.load_workbook(events_path)
        if "events" not in wb.sheetnames:
            log("Warning: 'events' sheet not found in events.xlsx")
            return True
            
        ws = wb["events"]
        
        # Find Alias and 掉落的道具ID columns
        alias_col = -1
        drop_id_col = -1
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(1, c).value or "").lower()
            if "alias" in h: alias_col = c
            if "掉落的道具id" in h: drop_id_col = c
            
        if alias_col == -1 or drop_id_col == -1:
            log(f"Warning: Could not find columns in events.xlsx (Alias={alias_col}, DropID={drop_id_col})")
            return True
            
        modified = False
        tgt_lower = tgt_evt.lower()
        for r in range(2, ws.max_row + 1):
            alias_val = str(ws.cell(r, alias_col).value or "").lower()
            if tgt_lower in alias_val:
                current_id_v = ws.cell(r, drop_id_col).value
                if current_id_v:
                    try:
                        cid = int(float(str(current_id_v)))
                        if cid in id_map:
                            new_id = id_map[cid]
                            ws.cell(r, drop_id_col, new_id)
                            log(f"events.xlsx: Updated drop ID for {tgt_evt}: {cid} -> {new_id}")
                            modified = True
                    except (ValueError, TypeError):
                        pass
        
        if modified:
            wb.save(events_path)
            log(f"events.xlsx successfully synchronized with new item IDs for {tgt_evt}.")
        return True
    except Exception as e:
        log(f"Error syncing events.xlsx: {e}")
        return False
            
    except Exception as e:
        log(f"Error updating item.xlsx: {e}")
        traceback.print_exc()
        return False

def update_event_shop_excel(proj_root, src_evt, tgt_evt, root_path, step_index):
    """Step 11: Update event_shop.xlsx with target shop and items."""
    import openpyxl
    import stat
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    shop_path = os.path.join(design_dir, "event_shop.xlsx")
    
    if not os.path.exists(shop_path):
        log(f"Error: event_shop.xlsx not found at {shop_path}")
        return False
    
    extra = load_extra_args(root_path)
    event_id = extra.get("event_id", "")
    log(f"event_shop: Received extra args: {extra}")
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()

    def swap_name(value):
        if value is None: return None
        s = str(value)
        if src_evt:
            s = s.replace(src_upper, tgt_upper)
            s = s.replace(src_lower, tgt_lower)
            s = s.replace(src_pascal, tgt_pascal)
            s = s.replace(src_evt, tgt_evt)
        # Fallback for ID-like columns (e.g. 20250701 -> 20260501)
        src_digits = "".join(filter(str.isdigit, src_evt))
        tgt_digits = "".join(filter(str.isdigit, tgt_evt))
        if len(src_digits) >= 4 and len(tgt_digits) >= 4:
            s = s.replace("20" + src_digits[:4], "20" + tgt_digits[:4])
        return s

    def get_id_value(val):
        """Extract numeric ID from potentially commented string '//100001'"""
        if val is None: return None
        s = str(val).strip()
        if s.startswith("//"):
            s = s[2:].strip()
        try:
            return int(s)
        except ValueError:
            return None

    try:
        # Backup
        backup_path = shop_path + ".bak"
        if os.path.exists(backup_path):
            os.chmod(backup_path, stat.S_IWRITE)
            os.remove(backup_path)
        shutil.copy2(shop_path, backup_path)
        record_file_creation(root_path, step_index, backup_path)
        
        wb = openpyxl.load_workbook(shop_path)
        modified = False
        
        # ========== Sheet 1: 物品 (Items) ==========
        item_sheet = "物品"
        if item_sheet in wb.sheetnames:
            ws_item = wb[item_sheet]
            real_max = get_real_max_row(ws_item)
            
            # 1. Find the highest ID and determine new group start
            max_id = 0
            for r in range(2, real_max + 1):
                id_val = get_id_value(ws_item.cell(r, 1).value)
                if id_val and id_val > max_id:
                    max_id = id_val
            
            new_group_start = (max_id // 100 + 1) * 100 + 1
            log(f"item_sheet: Last max ID was {max_id}. New group starts at {new_group_start}.")
            
            # 2. Find source items (those linked to src_evt)
            # We'll identify them by searching Alias
            src_item_rows = []
            for r in range(2, real_max + 1):
                alias_val = str(ws_item.cell(r, 2).value or "")
                if src_lower in alias_val.lower():
                    src_item_rows.append(r)
            
            if not src_item_rows:
                log(f"item_sheet: No items found for source {src_evt}. Searching by ID list from shop...")
                # Fallback: We'll read it from event_shop later, but let's try to find them now if possible.
            
            # 3. Clone source items
            new_item_ids = []
            if src_item_rows:
                current_id = new_group_start
                insert_pos = real_max + 1
                for src_r in src_item_rows:
                    row_data = [ws_item.cell(src_r, c).value for c in range(1, ws_item.max_column + 1)]
                    # Update ID (Column 1)
                    row_data[0] = current_id
                    # Swap names in ALL other columns (Column 2 onwards)
                    for i in range(1, len(row_data)):
                        if row_data[i] is not None:
                            row_data[i] = swap_name(row_data[i])
                    
                    # Write to new row
                    for c_idx, val in enumerate(row_data):
                        ws_item.cell(insert_pos, c_idx + 1, val)
                    
                    new_item_ids.append(current_id)
                    log(f"item_sheet: Added link item {current_id} (cloned from {src_r})")
                    current_id += 1
                    insert_pos += 1
                modified = True
            
            # 4. Commenting logic: only last two groups uncommented
            # A group is defined by its ID // 100 segment
            real_max = get_real_max_row(ws_item) # Refresh after adding
            # Find all group segments
            segments = []
            for r in range(2, real_max + 1):
                id_v = get_id_value(ws_item.cell(r, 1).value)
                if id_v:
                    seg = id_v // 100
                    if not segments or segments[-1] != seg:
                        segments.append(seg)
            
            if len(segments) > 2:
                keep_segments = segments[-2:]
                log(f"item_sheet: Keeping groups {keep_segments} active. Commenting others.")
                for r in range(2, real_max + 1):
                    val = ws_item.cell(r, 1).value
                    if val:
                        id_v = get_id_value(val)
                        if id_v and (id_v // 100) not in keep_segments:
                            s_val = str(val).strip()
                            if not s_val.startswith("//"):
                                ws_item.cell(r, 1, "//" + s_val)
                        elif id_v and (id_v // 100) in keep_segments:
                            s_val = str(val).strip()
                            if s_val.startswith("//"):
                                ws_item.cell(r, 1, s_val[2:].strip())
                modified = True

        # ========== Sheet 2: event_shop ==========
        shop_sheet = "event_shop"
        if shop_sheet in wb.sheetnames:
            ws_shop = wb[shop_sheet]
            real_max = get_real_max_row(ws_shop)
            
            # Find source row
            src_row = None
            for r in range(2, real_max + 1):
                val = str(ws_shop.cell(r, 2).value or "")
                if src_lower in val.lower():
                    src_row = r
                    break
            
            if src_row:
                new_row = real_max + 1
                for c in range(1, ws_shop.max_column + 1):
                    old_val = ws_shop.cell(src_row, c).value
                    ws_shop.cell(new_row, c, swap_name(old_val))
                
                # Update Shop ID
                if event_id:
                    ws_shop.cell(new_row, 1, int(event_id))
                
                # Update Item List (Col 4)
                if new_item_ids:
                    ws_shop.cell(new_row, 4, ",".join(map(str, new_item_ids)))
                
                log(f"event_shop: Added shop row {new_row} for {tgt_evt}.")
                
                # Commenting logic: only last two rows uncommented
                # After adding, new_row is the latest
                for r in range(2, real_max + 2):
                    val = ws_shop.cell(r, 1).value
                    if val:
                        s_val = str(val).strip()
                        if r < (real_max + 2 - 2): # All but last two
                            if not s_val.startswith("//"):
                                ws_shop.cell(r, 1, "//" + s_val)
                        else: # Last two
                            if s_val.startswith("//"):
                                ws_shop.cell(r, 1, s_val[2:].strip())
                
                modified = True
            else:
                log(f"event_shop: Source event {src_evt} not found. Skipping shop creation.")

        if modified:
            os.chmod(shop_path, stat.S_IWRITE | stat.S_IREAD)
            wb.save(shop_path)
            log("event_shop.xlsx saved successfully.")
        else:
            log("event_shop.xlsx: No changes were needed.")
            
        return True
    except Exception as e:
        log(f"Exception updating event_shop.xlsx: {e}")
        traceback.print_exc()
        return False

def update_icon_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 13: Update icon.xlsx and intelligently auto-generate missing icon hooks."""
    import openpyxl
    import stat
    import os
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    icon_path = os.path.join(design_dir, "icon.xlsx")
    item_path = os.path.join(design_dir, "item.xlsx")
    
    if not os.path.exists(icon_path):
        log(f"Error: icon.xlsx not found at {icon_path}")
        return False
        
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()

    def swap_name(value):
        if value is None: return None
        s = str(value)
        if src_evt:
            s = s.replace(src_upper, tgt_upper)
            s = s.replace(src_lower, tgt_lower)
            s = s.replace(src_pascal, tgt_pascal)
            s = s.replace(src_evt, tgt_evt)
        return s

    # 1. First, gather REQUIRED target icons from item.xlsx AND the target event excel
    required_tgt_icons = set()
    try:
        # A. From item.xlsx
        if os.path.exists(item_path):
            wb_item = openpyxl.load_workbook(item_path, data_only=True)
            if "item" in wb_item.sheetnames:
                ws_item = wb_item["item"]
                for r in range(2, ws_item.max_row + 1):
                    alias = str(ws_item.cell(r, 2).value or "")
                    name = str(ws_item.cell(r, 5).value or "")
                    if tgt_evt.lower() in alias.lower() or tgt_evt.lower() in name.lower():
                        icon7 = str(ws_item.cell(r, 7).value or "").strip()
                        icon8 = str(ws_item.cell(r, 8).value or "").strip()
                        icon9 = str(ws_item.cell(r, 9).value or "").strip()
                        if icon7: required_tgt_icons.add(icon7)
                        if icon8: required_tgt_icons.add(icon8)
                        if icon9: required_tgt_icons.add(icon9)
            wb_item.close()
            
        # B. From tgt_evt excel (e.g. sim2605.xlsx)
        tgt_evt_path = os.path.join(design_dir, f"{tgt_evt}.xlsx")
        if not os.path.exists(tgt_evt_path):
            # Try lowercase
            tgt_evt_path = os.path.join(design_dir, f"{tgt_evt.lower()}.xlsx")
            
        if os.path.exists(tgt_evt_path):
            wb_tgt_evt = openpyxl.load_workbook(tgt_evt_path, data_only=True)
            for ws_t in wb_tgt_evt.worksheets:
                for row in ws_t.iter_rows():
                    for cell in row:
                        val = str(cell.value or "")
                        # Match icons (e.g. SIM2605_DISH_ICON_1)
                        if "_ICON" in val.upper() and tgt_upper in val.upper():
                            # Clean up (sometimes values have suffixes or are part of strings)
                            match = re.search(rf'({tgt_upper}_[A-Za-z0-9_]+_ICON_[0-9]+)', val, re.IGNORECASE)
                            if match:
                                required_tgt_icons.add(match.group(1).upper())
                            elif "_" in val:
                                 # Fallback: if it has _ICON, add it if it looks like an identifier
                                 if re.match(r'^[A-Za-z0-9_]+$', val):
                                     required_tgt_icons.add(val.strip().upper())
            wb_tgt_evt.close()
            log(f"icon_sheet: Identified {len(required_tgt_icons)} required icons from item.xlsx and {tgt_evt}.xlsx")
            
    except Exception as e:
        log(f"icon_sheet: Failed to gather required icons for cross-validation: {e}")

    try:
        os.chmod(icon_path, stat.S_IWRITE | stat.S_IREAD)
        wb_icon = openpyxl.load_workbook(icon_path)
        
        modified = False
        gui_event_id = None
        try:
            extra_args_path = os.path.join(log_dir, "extra_args.json")
            if os.path.exists(extra_args_path):
                with open(extra_args_path, "r") as f:
                    extra = json.load(f)
                    gui_event_id = str(extra.get("event_id", ""))
        except: pass

        # 2a. Handle Specialized icon_item templates (Rows 69 & 70)
        if "icon_item" in wb_icon.sheetnames:
            ws_item = wb_icon["icon_item"]
            
            # First, cleanup any previously mis-appended rows (due to logic bugs)
            # Find the BP divider if it exists
            divider_idx = ws_item.max_row + 1
            for r in range(1, ws_item.max_row + 1):
                if "// BP相关" in str(ws_item.cell(r, 1).value or ""):
                    divider_idx = r
                    break
            
            # If the divider exists, remove any cloned icons that were accidentally appended AFTER it
            if divider_idx < ws_item.max_row:
                for r in range(ws_item.max_row, divider_idx, -1):
                    val = str(ws_item.cell(r, 1).value or "").upper()
                    if tgt_upper in val and ("_TOKEN" in val or "_MATERIAL" in val):
                        ws_item.delete_rows(r)
                        log(f"icon_sheet: Cleaned up mis-appended row {r} from icon_item tail")
                        modified = True

            # Templates for TOKEN and MATERIAL
            for r_idx in [69, 70]:
                orig_key = str(ws_item.cell(r_idx, 1).value or "")
                if not orig_key or "_" not in orig_key: continue
                
                parts = orig_key.split('_', 1)
                orig_prefix = parts[0]
                new_key = f"{tgt_upper}_{parts[1]}"
                
                already_exists = False
                for r in range(1, ws_item.max_row + 1):
                    if str(ws_item.cell(r, 1).value).upper() == str(new_key).upper():
                        already_exists = True
                        break
                
                if not already_exists:
                    # Insert at the divider to keep it grouped correctly
                    ws_item.insert_rows(divider_idx)
                    new_r = divider_idx
                    for c in range(1, ws_item.max_column + 1):
                        v = ws_item.cell(r_idx, c).value
                        if v:
                            nv = swap_name(v)
                            if orig_prefix.upper() in str(nv).upper():
                                nv = nv.replace(orig_prefix.upper(), tgt_upper)
                                nv = nv.replace(orig_prefix.lower(), tgt_lower)
                            if c == 3 and nv and gui_event_id:
                                nv = re.sub(r'\d{8}', gui_event_id, str(nv))
                            ws_item.cell(new_r, c, nv)
                        else:
                            ws_item.cell(new_r, c, None)
                    log(f"icon_sheet: Inserted template row {r_idx} ({new_key}) into icon_item at Row {new_r}")
                    divider_idx += 1 # Advance divider index for the next insertion
                    modified = True

        # 2b. Extract from backup and transfer to active sheets
        backup_sheet_name = "icon_备份"
        active_sheet_name = "icon_event"
        
        if backup_sheet_name in wb_icon.sheetnames and active_sheet_name in wb_icon.sheetnames:
            ws_backup = wb_icon[backup_sheet_name]
            ws_active = wb_icon[active_sheet_name]
            
            existing_active_icons = set()
            for r in range(2, ws_active.max_row + 1):
                icon_val = str(ws_active.cell(r, 1).value or "").strip()
                if icon_val: existing_active_icons.add(icon_val.upper())

            src_rows = []
            for r in range(2, ws_backup.max_row + 1):
                c1 = str(ws_backup.cell(r, 1).value or "").lower()
                c2 = str(ws_backup.cell(r, 2).value or "").lower()
                c3 = str(ws_backup.cell(r, 3).value or "").lower()
                if src_lower in c1 or src_lower in c2 or src_lower in c3:
                    src_rows.append(r)
            
            if src_rows:
                divider_row = ws_active.max_row + 1
                ws_active.cell(divider_row, 1, f"// {tgt_upper} (CLONED FROM {src_upper})")
                
                cloned_count = 0
                for r_src in src_rows:
                    new_icon_name = swap_name(ws_backup.cell(r_src, 1).value)
                    if new_icon_name and str(new_icon_name).upper().strip() in existing_active_icons:
                        continue
                        
                    new_row_idx = ws_active.max_row + 1
                    for c in range(1, ws_backup.max_column + 1):
                        old_val = ws_backup.cell(r_src, c).value
                        new_val = swap_name(old_val)
                        # Sync Event ID in Sprite Name (Col 3) as requested
                        if c == 3 and new_val and gui_event_id:
                            new_val = re.sub(r'\d{8}', gui_event_id, str(new_val))
                        ws_active.cell(new_row_idx, c, new_val)
                    cloned_count += 1
                
                if cloned_count > 0:
                    log(f"icon_sheet: Cloned {cloned_count} rows from backup to active event sheet.")
                    modified = True

        # 3. Traditional sheet-by-sheet update
        for sheet_name in wb_icon.sheetnames:
            if sheet_name in [backup_sheet_name, active_sheet_name, "icon_item"]:
                continue
                
            ws = wb_icon[sheet_name]
            src_rows = []
            for r in range(2, ws.max_row + 1):
                c1 = str(ws.cell(r, 1).value or "").lower()
                c2 = str(ws.cell(r, 2).value or "").lower()
                c3 = str(ws.cell(r, 3).value or "").lower()
                if src_lower in c1 or src_lower in c2 or src_lower in c3:
                    src_rows.append(r)
            
            if src_rows:
                existing_in_sheet = set()
                for r in range(2, ws.max_row + 1):
                    v = str(ws.cell(r, 1).value or "").strip().upper()
                    if v: existing_in_sheet.add(v)

                divider_row = ws.max_row + 1
                ws.cell(divider_row, 1, f"// {tgt_upper}")
                
                cloned_in_sheet = 0
                for r_src in src_rows:
                    new_key = swap_name(ws.cell(r_src, 1).value)
                    if new_key and str(new_key).strip().upper() in existing_in_sheet:
                        continue
                    new_row_idx = ws.max_row + 1
                    for c in range(1, ws.max_column + 1):
                        old_v = ws.cell(r_src, c).value
                        new_v = swap_name(old_v)
                        # Sync Event ID in Sprite Name (Col 3)
                        if c == 3 and new_v and gui_event_id:
                            new_v = re.sub(r'\d{8}', gui_event_id, str(new_v))
                        ws.cell(new_row_idx, c, new_v)
                    cloned_in_sheet += 1
                
                if cloned_in_sheet > 0:
                    log(f"icon_sheet: Cloned {cloned_in_sheet} rows in [{sheet_name}] for {tgt_evt}.")
                    modified = True

        # 4. Intelligent Missing Icon Generation
        all_final_icons = set()
        for sheet_name in wb_icon.sheetnames:
            if "备份" in sheet_name: continue
            ws_f = wb_icon[sheet_name]
            for r in range(2, ws_f.max_row + 1):
                v = str(ws_f.cell(r, 1).value or "").strip().upper()
                if v: all_final_icons.add(v)
                
        missing_icons = []
        for req_icon in required_tgt_icons:
            if req_icon.upper() not in all_final_icons:
                missing_icons.append(req_icon)

        if missing_icons:
            ws_target_fix = wb_icon["icon_event"] if "icon_event" in wb_icon.sheetnames else wb_icon.worksheets[0]
            source_to_clone = {}
            for sheet_name in wb_icon.sheetnames:
                ws_search = wb_icon[sheet_name]
                for r in range(2, ws_search.max_row + 1):
                    s1 = str(ws_search.cell(r, 1).value or "").strip().upper()
                    if s1 and s1 not in source_to_clone:
                        source_to_clone[s1] = (ws_search, r)
            
            divider_row = ws_target_fix.max_row + 1
            ws_target_fix.cell(divider_row, 1, f"// {tgt_upper} (INTELLIGENT RECOVERY)")
            
            for m_icon in missing_icons:
                src_icon_name = m_icon.replace(tgt_upper, src_upper).replace(tgt_lower, src_lower).replace(tgt_pascal, src_pascal)
                if src_icon_name.upper() in source_to_clone:
                    ws_src, r_idx = source_to_clone[src_icon_name.upper()]
                    new_r = ws_target_fix.max_row + 1
                    for c in range(1, ws_src.max_column + 1):
                        old_v = ws_src.cell(r_idx, c).value
                        new_v = swap_name(old_v)
                        if c == 3 and new_v and gui_event_id:
                            new_v = re.sub(r'\d{8}', gui_event_id, str(new_v))
                        ws_target_fix.cell(new_r, c, new_v)
                    log(f"icon_sheet: Successfully RECOVERED icon '{m_icon}' from {ws_src.title}")
                else:
                    new_r = ws_target_fix.max_row + 1
                    ws_target_fix.cell(new_r, 1, m_icon)
                    ws_target_fix.cell(new_r, 2, f"{m_icon}")
                    ws_target_fix.cell(new_r, 3, f"{gui_event_id}_TBD" if gui_event_id else "Sprite_TBD")
                    ws_target_fix.cell(new_r, 4, "Atlas_Quest")
                    log(f"WARNING: Icon '{m_icon}' missing. Created placeholder in {ws_target_fix.title}.")
            modified = True
        
        if modified:
            wb_icon.save(icon_path)
            log(f"icon.xlsx successfully updated for {tgt_evt}.")
            return True
        else:
            log(f"icon.xlsx: No changes were needed.")
            return True
    
    except Exception as e:
        log(f"Error updating icon.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False

    except Exception as e:
        log(f"Error updating icon.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_quiz_date_prefix(proj_root, event_name):
    """Derive quiz date prefix (YYMMDD) from events.xlsx start time for a given event."""
    import openpyxl
    import os
    
    events_path = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design\events.xlsx")
    if not os.path.exists(events_path):
        log(f"Warning: events.xlsx not found at {events_path}")
        return None
    
    wb = openpyxl.load_workbook(events_path, data_only=True)
    ws = wb["events"]
    
    # Find Alias column and 开始时间 column
    alias_col = None
    start_time_col = None
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(1, c).value or "")
        if "Alias" in header or "alias" in header.lower():
            alias_col = c
        if "开始时间" in header:
            start_time_col = c
    
    if not alias_col or not start_time_col:
        log(f"Warning: Could not find Alias or 开始时间 column in events.xlsx")
        return None
    
    evt_lower = event_name.lower()
    for r in range(2, ws.max_row + 1):
        alias_val = str(ws.cell(r, alias_col).value or "").lower()  # type: ignore
        if evt_lower in alias_val:
            start_val = ws.cell(r, start_time_col).value  # type: ignore
            if start_val:
                # Could be datetime object or string like "2025-07-17 00:00:00"
                from datetime import datetime
                if isinstance(start_val, datetime):
                    return start_val.strftime("%y%m%d")
                else:
                    s = str(start_val).strip()
                    # Parse "2025-07-17 00:00:00" or similar
                    try:
                        dt = datetime.strptime(s[:10], "%Y-%m-%d")
                        return dt.strftime("%y%m%d")
                    except Exception:
                        log(f"Warning: Could not parse start time '{s}' for {event_name}")
                        return None
    
    log(f"Warning: Could not find event '{event_name}' in events.xlsx")
    return None


def update_localization_main_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 14 Part A: Update localization.xlsx (main localization file only)."""
    import openpyxl
    import stat
    import os
    import re
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "localization.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Warning: localization.xlsx not found. Skipping.")
        return True
    
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()

    src_num_match = re.search(r'\d+', src_evt)
    tgt_num_match = re.search(r'\d+', tgt_evt)
    src_digits = src_num_match.group() if src_num_match else ""
    tgt_digits = tgt_num_match.group() if tgt_num_match else ""

    # 1. First, gather REQUIRED target keys from the target event excel
    required_tgt_keys = set()
    try:
        tgt_evt_path = os.path.join(design_dir, f"{tgt_evt}.xlsx")
        if not os.path.exists(tgt_evt_path):
            tgt_evt_path = os.path.join(design_dir, f"{tgt_evt.lower()}.xlsx")
            
        if os.path.exists(tgt_evt_path):
            wb_tgt_evt = openpyxl.load_workbook(tgt_evt_path, data_only=True)
            for ws_t in wb_tgt_evt.worksheets:
                for row in ws_t.iter_rows():
                    for cell in row:
                        val = str(cell.value or "")
                        # Match localization keys: e.g. SIM2605_ROOMNAME_1 or LOC_SIM2605_...
                        # Usually uppercase identifiers starting with SIM or having the event digits
                        if (tgt_upper in val.upper() or (tgt_digits and tgt_digits in val)) and "_" in val:
                            match = re.search(rf'([A-Z0-9_]*{tgt_upper}[A-Z0-9_]*)', val.upper())
                            if match:
                                required_tgt_keys.add(match.group(1))
                            elif re.match(r'^[A-Z0-9_]+$', val.upper()):
                                required_tgt_keys.add(val.upper().strip())
            wb_tgt_evt.close()
            log(f"localization: Identified {len(required_tgt_keys)} required keys from {tgt_evt}.xlsx")
            
    except Exception as e:
        log(f"localization: Failed to gather required keys from {tgt_evt}.xlsx: {e}")

    def swap_name(value):
        if value is None: return None
        s = str(value)  # type: ignore
        if src_evt:
            s = s.replace(src_upper, tgt_upper)
            s = s.replace(src_lower, tgt_lower)
            s = s.replace(src_pascal, tgt_pascal)
            s = s.replace(src_evt, tgt_evt)
        if src_digits and tgt_digits:
            s = s.replace(src_digits, tgt_digits)
        return s

    try:
        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        wb = openpyxl.load_workbook(file_path)
        
        modified = False
        for sheet_name in wb.sheetnames:
            if sheet_name in ["languages", "country", "localize_零使用"]:
                continue
                
            ws = wb[sheet_name]
            src_rows = []
            
            # Map of key -> row_idx in this sheet to prevent duplicates
            existing_keys = {}
            for r in range(2, ws.max_row + 1):
                k = str(ws.cell(r, 1).value or "").upper().strip()
                if k: existing_keys[k] = r
            
            # Find rows to clone (fuzzy match on name OR digits)
            for r in range(2, ws.max_row + 1):
                c1 = str(ws.cell(r, 1).value or "").lower()
                c2 = str(ws.cell(r, 2).value or "").lower()
                
                match_found = False
                if src_lower in c1 or src_lower in c2:
                    match_found = True
                elif src_digits and (src_digits in c1 or src_digits in c2):
                    # More cautious match for digits: must look like an event-prefix
                    if re.search(rf'([A-Za-z]+{src_digits}|{src_digits}_)', c1):
                        match_found = True
                
                if match_found:
                    src_rows.append(r)
            
            if src_rows:
                # Add a divider
                new_divider = ws.max_row + 1
                ws.cell(new_divider, 1, f"// {tgt_upper} AUTO-CLONED")
                
                cloned_in_sheet = 0
                for r in src_rows:
                    new_key = swap_name(ws.cell(r, 1).value)
                    if new_key and str(new_key).upper().strip() in existing_keys:
                        continue # Skip if already exists
                        
                    new_row_idx = ws.max_row + 1
                    for c in range(1, ws.max_column + 1):
                        ws.cell(new_row_idx, c, swap_name(ws.cell(r, c).value))
                    cloned_in_sheet += 1
                
                if cloned_in_sheet > 0:
                    log(f"localization: Cloned {cloned_in_sheet} rows in [{sheet_name}] for {tgt_evt}")
                    modified = True
        
        # 2. Final verification against REQUIRED keys from target excel
        # Just in case some keys were missed in the cloning phase
        missing_keys = []
        # Pre-scan EVERYTHING again to be sure
        all_final_keys = set()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                all_final_keys.add(str(ws.cell(r, 1).value or "").upper().strip())
        
        for r_key in required_tgt_keys:
            if r_key not in all_final_keys:
                missing_keys.append(r_key)
        
        if missing_keys:
            # Try to recover missing keys from ANY sheet of localization.xlsx
            # (e.g. if SIM2512_ROOMNAME_1 was in a different sheet than expected)
            ws_fix = wb[wb.sheetnames[0]] # Just put in first sheet or a common one
            if "localize_room" in wb.sheetnames: ws_fix = wb["localize_room"]
            
            recovery_count = 0
            # Build global source map for recovery
            source_map = {}
            for sheet_name in wb.sheetnames:
                ws_s = wb[sheet_name]
                for r in range(2, ws_s.max_row + 1):
                    k_val = str(ws_s.cell(r, 1).value or "")
                    if k_val: source_map[k_val.upper()] = [ws_s.cell(r, c).value for c in range(1, ws_s.max_column + 1)]
            
            new_divider = ws_fix.max_row + 1
            ws_fix.cell(new_divider, 1, f"// {tgt_upper} INTELLIGENT RECOVERY")
            
            for m_key in missing_keys:
                src_key = m_key.replace(tgt_upper, src_upper).replace(tgt_digits, src_digits)
                if src_key in source_map:
                    new_r = ws_fix.max_row + 1
                    row_data = source_map[src_key]
                    for c_idx, val in enumerate(row_data):
                        ws_fix.cell(new_r, c_idx + 1, swap_name(val))
                    recovery_count += 1
                    log(f"localization: Recovered missing key '{m_key}' from source '{src_key}'")
                else:
                    # Pure placeholder
                    new_r = ws_fix.max_row + 1
                    ws_fix.cell(new_r, 1, m_key)
                    ws_fix.cell(new_r, 2, f"[FIX ME] {m_key}")
                    log(f"WARNING: Localization key '{m_key}' is REQUIRED by {tgt_evt}.xlsx but was NOT FOUND in localization.xlsx. Created placeholder.")
            
            if len(missing_keys) > 0:
                modified = True
            for r in range(2, ws.max_row + 1):
                val1 = str(ws.cell(r, 1).value or "").lower()  # type: ignore
                val2 = str(ws.cell(r, 2).value or "").lower()  # type: ignore
                
                match = False
                if src_lower and (src_lower in val1 or src_lower in val2):
                    match = True
                elif src_digits and (src_digits in val1 or src_digits in val2):
                    match = True
                    
                if match:
                    src_rows.append(r)
                    
            if src_rows:
                divider_row = ws.max_row + 1
                ws.cell(divider_row, 1, f"// {tgt_upper}")
                
                for r in src_rows:
                    new_row_idx = ws.max_row + 1
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(r, c).value  # type: ignore
                        ws.cell(new_row_idx, c, swap_name(old_val))  # type: ignore
                        
                log(f"localization.xlsx [{sheet_name}]: Cloned {len(src_rows)} rows for {tgt_evt}.")
                modified = True
                
        if modified:
            wb.save(file_path)
            log(f"localization.xlsx successfully updated for {tgt_evt}.")
        else:
            log(f"localization.xlsx: No changes needed. No source strings found.")
            
    except Exception as e:
        log(f"Error updating localization.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False
        
    return True


def update_localization_quiz_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 14 Part B: Update localization_quiz.xlsx with backup/clone workflow."""
    import openpyxl
    import stat
    import os
    import json
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "localization_quiz.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Warning: localization_quiz.xlsx not found. Skipping.")
        return True
    
    # --- Derive quiz date prefixes ---
    src_quiz_date = get_quiz_date_prefix(proj_root, src_evt)
    if not src_quiz_date:
        log(f"Warning: Could not derive source quiz date for {src_evt}. Checking extra_args...")
        return True  # Non-fatal: quiz may not apply to all events
    
    # Primary: derive from events.xlsx (Step 10 already populated correct dates)
    tgt_quiz_date = get_quiz_date_prefix(proj_root, tgt_evt)
    
    # Fallback: check extra_args.json start_time if events.xlsx doesn't have target yet
    if not tgt_quiz_date:
        extra_args_path = os.path.join(log_dir, "extra_args.json")
        if os.path.exists(extra_args_path):
            try:
                with open(extra_args_path, "r") as f:
                    extra = json.load(f)
                start_time_str = extra.get("start_time", "")
                if start_time_str:
                    from datetime import datetime
                    try:
                        dt = datetime.strptime(start_time_str[:10], "%Y-%m-%d")
                        tgt_quiz_date = dt.strftime("%y%m%d")
                    except Exception:
                        pass
            except Exception:
                pass
    
    if not tgt_quiz_date:
        log(f"Warning: Could not derive target quiz date for {tgt_evt}. Cannot process localization_quiz.xlsx.")
        return True  # Non-fatal
    
    src_quiz_prefix = f"QUIZ{src_quiz_date}"  # e.g., QUIZ250717
    tgt_quiz_prefix = f"QUIZ{tgt_quiz_date}"  # e.g., QUIZ260501
    
    log(f"Quiz date derivation: {src_evt} -> {src_quiz_prefix}, {tgt_evt} -> {tgt_quiz_prefix}")
    
    def swap_quiz(value):
        """Swap quiz date prefixes in cell values."""
        if value is None: return None
        s = str(value)  # type: ignore
        s = s.replace(src_quiz_prefix.upper(), tgt_quiz_prefix.upper())
        s = s.replace(src_quiz_prefix.lower(), tgt_quiz_prefix.lower())
        # Also handle mixed case
        s = s.replace(src_quiz_prefix, tgt_quiz_prefix)
        return s
    
    try:
        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        wb = openpyxl.load_workbook(file_path)
        
        quiz_sheet_name = "Quiz"
        backup_sheet_name = "Quiz_备份"
        
        if quiz_sheet_name not in wb.sheetnames or backup_sheet_name not in wb.sheetnames:
            log(f"Warning: Required sheets '{quiz_sheet_name}' or '{backup_sheet_name}' not found.")
            return True
        
        ws_quiz = wb[quiz_sheet_name]
        ws_backup = wb[backup_sheet_name]
        
        # --- Phase 1: Archive current Quiz rows to Quiz_备份 ---
        # Collect existing backup keys for deduplication
        existing_backup_keys = set()
        for r in range(2, ws_backup.max_row + 1):
            key = str(ws_backup.cell(r, 1).value or "").strip()  # type: ignore
            if key:
                existing_backup_keys.add(key.lower())
        
        archived_count = 0
        for r in range(2, ws_quiz.max_row + 1):
            quiz_key = str(ws_quiz.cell(r, 1).value or "").strip()  # type: ignore
            if not quiz_key:
                continue
            
            # Only archive if not already in backup
            if quiz_key.lower() not in existing_backup_keys:
                new_backup_row = ws_backup.max_row + 1
                for c in range(1, ws_quiz.max_column + 1):
                    ws_backup.cell(new_backup_row, c, ws_quiz.cell(r, c).value)  # type: ignore
                archived_count += 1
        
        log(f"localization_quiz: Archived {archived_count} new rows from Quiz to Quiz_备份.")
        
        # --- Phase 2: Clear Quiz sheet (keep row 1 header) ---
        # Save headers first
        headers = []
        for c in range(1, ws_quiz.max_column + 1):
            headers.append(ws_quiz.cell(1, c).value)  # type: ignore
        
        # Delete all data rows
        if ws_quiz.max_row > 1:
            ws_quiz.delete_rows(2, ws_quiz.max_row - 1)
        
        log(f"localization_quiz: Cleared Quiz sheet (kept headers).")
        
        # --- Phase 3: Clone source quiz rows from backup ---
        src_prefix_lower = src_quiz_prefix.lower()
        cloned_count = 0
        
        for r in range(2, ws_backup.max_row + 1):
            val = str(ws_backup.cell(r, 1).value or "").lower()  # type: ignore
            if src_prefix_lower in val:
                new_row = ws_quiz.max_row + 1
                for c in range(1, ws_backup.max_column + 1):
                    old_val = ws_backup.cell(r, c).value  # type: ignore
                    ws_quiz.cell(new_row, c, swap_quiz(old_val))  # type: ignore
                cloned_count += 1
        
        log(f"localization_quiz: Cloned {cloned_count} rows from Quiz_备份 ({src_quiz_prefix} -> {tgt_quiz_prefix}) into Quiz sheet.")
        
        wb.save(file_path)
        log(f"localization_quiz.xlsx successfully updated.")
        return True
        
    except Exception as e:
        log(f"Error updating localization_quiz.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_answer_challenge_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 14 Part C: Update answer_challenge.xlsx (list, questions, stage rewards)."""
    import openpyxl
    import stat
    import os
    import re
    import json
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "answer_challenge.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Warning: answer_challenge.xlsx not found. Skipping.")
        return True
    
    # --- Get quiz date prefixes ---
    src_quiz_date = get_quiz_date_prefix(proj_root, src_evt)
    # Primary: derive from events.xlsx (Step 10 already populated correct dates)
    tgt_quiz_date = get_quiz_date_prefix(proj_root, tgt_evt)
    
    # Fallback: check extra_args.json start_time if events.xlsx doesn't have target yet
    if not tgt_quiz_date:
        extra_args_path = os.path.join(log_dir, "extra_args.json")
        if os.path.exists(extra_args_path):
            try:
                with open(extra_args_path, "r") as f:
                    extra = json.load(f)
                start_time_str = extra.get("start_time", "")
                if start_time_str:
                    from datetime import datetime
                    try:
                        dt = datetime.strptime(start_time_str[:10], "%Y-%m-%d")
                        tgt_quiz_date = dt.strftime("%y%m%d")
                    except Exception:
                        pass
            except Exception:
                pass
    
    if not tgt_quiz_date or not src_quiz_date:
        log(f"Warning: Could not derive quiz dates for answer_challenge update. Skipping.")
        return True
    
    src_quiz_prefix = f"QUIZ{src_quiz_date}"
    tgt_quiz_prefix = f"QUIZ{tgt_quiz_date}"
    
    # Swap helpers for event names in stage rewards
    src_lower = src_evt.lower()
    tgt_lower = tgt_evt.lower()
    src_upper = src_evt.upper()
    tgt_upper = tgt_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt
    src_num = re.search(r'\d+', src_evt)
    tgt_num = re.search(r'\d+', tgt_evt)
    src_digits = src_num.group() if src_num else ""
    tgt_digits = tgt_num.group() if tgt_num else ""
    
    def swap_name(value):
        if value is None: return None
        s = str(value)  # type: ignore
        # Swap quiz prefixes
        s = s.replace(src_quiz_prefix.upper(), tgt_quiz_prefix.upper())
        s = s.replace(src_quiz_prefix.lower(), tgt_quiz_prefix.lower())
        s = s.replace(src_quiz_prefix, tgt_quiz_prefix)
        # Swap event names
        if src_evt:
            s = s.replace(src_upper, tgt_upper)
            s = s.replace(src_lower, tgt_lower)
            s = s.replace(src_pascal, tgt_pascal)
            s = s.replace(src_evt, tgt_evt)
        # ONLY swap digits if they aren't already part of the target event name
        # to prevent "Sim26052605" errors
        if src_digits and tgt_digits and src_digits not in tgt_evt:
            s = s.replace(src_digits, tgt_digits)
        return s
    
    try:
        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        # === Part C1: answer_challenge_list — add new row ===
        if "answer_challenge_list" in wb.sheetnames:
            ws_list = wb["answer_challenge_list"]
            
            # Find last data row and auto-increment ID
            last_id = 0
            last_row = 1
            for r in range(2, ws_list.max_row + 1):
                id_val = ws_list.cell(r, 1).value  # type: ignore
                if id_val is not None:
                    try:
                        last_id = int(id_val)
                        last_row = r
                    except (ValueError, TypeError):
                        pass
            
            # Check for idempotency: if an alias for this target event is already on the list, skip appending!
            already_exists = False
            for r in range(2, ws_list.max_row + 1):
                existing_alias = str(ws_list.cell(r, 2).value or "").lower()
                if tgt_evt.lower() in existing_alias:
                    already_exists = True
                    break
            
            if already_exists:
                log(f"answer_challenge [answer_challenge_list]: Target event {tgt_evt} already exists. Skipping row generation.")
            else:
                new_id = last_id + 1
                new_row = last_row + 1
            
            # Copy the last row as template and swap names
            for c in range(1, ws_list.max_column + 1):
                old_val = ws_list.cell(last_row, c).value  # type: ignore
                ws_list.cell(new_row, c, swap_name(old_val))  # type: ignore
            
            # Override the ID
            ws_list.cell(new_row, 1, new_id)  # type: ignore
            
            # Ensure Alias is unique by appending target event if it doesn't already contain it
            current_alias = str(ws_list.cell(new_row, 2).value or "")
            if tgt_evt.lower() not in current_alias.lower():
                ws_list.cell(new_row, 2, f"{current_alias}_{tgt_evt}")
                
            # Override Event ID and Times using extra_args
            extra_args_path = os.path.join(log_dir, "extra_args.json")
            if os.path.exists(extra_args_path):
                try:
                    with open(extra_args_path, "r") as f:
                        extra = json.load(f)
                    start_time_str = extra.get("start_time", "")
                    end_time_str = extra.get("end_time", "")
                    event_id_str = extra.get("event_id", "")
                    
                    if start_time_str: ws_list.cell(new_row, 4, start_time_str)
                    if end_time_str: ws_list.cell(new_row, 5, end_time_str)
                    if event_id_str: ws_list.cell(new_row, 6, int(event_id_str))
                except Exception as e:
                    log(f"Warning: Failed to inject extra args into answer_challenge_list: {e}")
            
                log(f"answer_challenge [answer_challenge_list]: Added row {new_row} with ID={new_id}")
                modified = True
        
        # === Part C2: 题目 (Questions) — auto-detect current quiz prefix and replace ===
        if "题目" in wb.sheetnames:
            ws_q = wb["题目"]
            q_modified = 0
            
            # Auto-detect the current quiz prefix from first data row
            current_prefix = None
            for c in range(1, ws_q.max_column + 1):
                cell_val = str(ws_q.cell(2, c).value or "")  # type: ignore
                prefix_match = re.search(r'(QUIZ\d{6})', cell_val, re.IGNORECASE)
                if prefix_match:
                    current_prefix = prefix_match.group(1).upper()
                    break
            
            if current_prefix and current_prefix != tgt_quiz_prefix:
                log(f"answer_challenge [题目]: Detected current prefix '{current_prefix}', replacing with '{tgt_quiz_prefix}'")
                for r in range(2, ws_q.max_row + 1):
                    for c in range(1, ws_q.max_column + 1):
                        cell_val = ws_q.cell(r, c).value  # type: ignore
                        if cell_val is not None:
                            s = str(cell_val)
                            if current_prefix in s or current_prefix.lower() in s.lower():
                                new_s = s.replace(current_prefix, tgt_quiz_prefix)
                                new_s = new_s.replace(current_prefix.lower(), tgt_quiz_prefix.lower())
                                ws_q.cell(r, c, new_s)  # type: ignore
                                q_modified += 1
                
                if q_modified > 0:
                    log(f"answer_challenge [题目]: Updated {q_modified} quiz key references ({current_prefix} -> {tgt_quiz_prefix})")
                    modified = True
            else:
                log(f"answer_challenge [题目]: No prefix change needed (current: {current_prefix}, target: {tgt_quiz_prefix})")
        
        # === Part C3: stage (Rewards) — auto-detect event reward references and replace ===
        if "stage" in wb.sheetnames:
            ws_stage = wb["stage"]
            stage_modified = 0
            
            # Auto-detect the current event name in stage rewards
            # Pattern: {EventName}{4digits}能量 or {EVENTNAME}{4digits}代币
            detected_event_name = None
            detected_event_digits = None
            for r in range(2, ws_stage.max_row + 1):
                for c in range(1, ws_stage.max_column + 1):
                    cell_val = str(ws_stage.cell(r, c).value or "")  # type: ignore
                    # Match patterns like "Merge2512能量" or "MERGE2512代币"
                    evt_match = re.search(r'([A-Za-z]+)(\d{4})(能量|代币)', cell_val)
                    if evt_match:
                        detected_event_name = evt_match.group(1)
                        detected_event_digits = evt_match.group(2)
                        break
                if detected_event_name:
                    break
            
            if detected_event_name and detected_event_digits:
                det_lower = detected_event_name.lower()
                det_upper = detected_event_name.upper()
                det_pascal = detected_event_name[0].upper() + detected_event_name[1:]
                
                log(f"answer_challenge [stage]: Detected '{detected_event_name}{detected_event_digits}' in rewards, replacing with '{tgt_pascal}{tgt_digits}'")
                
                for r in range(2, ws_stage.max_row + 1):
                    for c in range(1, ws_stage.max_column + 1):
                        cell_val = ws_stage.cell(r, c).value  # type: ignore
                        if cell_val is not None:
                            s = str(cell_val)
                            new_s = s
                            # Replace all case variants: Merge2512 -> Sim2605, MERGE2512 -> SIM2605
                            # Fixed replacement to avoid redundant digit concatenation
                            # If tgt_upper already includes tgt_digits (e.g. SIM2605), don't append it again.
                            new_s = new_s.replace(f"{det_upper}{detected_event_digits}", tgt_upper)
                            new_s = new_s.replace(f"{det_lower}{detected_event_digits}", tgt_lower)
                            new_s = new_s.replace(f"{det_pascal}{detected_event_digits}", tgt_pascal)
                            if new_s != s:
                                ws_stage.cell(r, c, new_s)  # type: ignore
                                stage_modified += 1
                
                if stage_modified > 0:
                    log(f"answer_challenge [stage]: Updated {stage_modified} reward references")
                    modified = True
            else:
                log(f"answer_challenge [stage]: No event reward references detected.")
        
        if modified:
            wb.save(file_path)
            log(f"answer_challenge.xlsx successfully updated.")
        else:
            log(f"answer_challenge.xlsx: No changes needed.")
        
        return True
        
    except Exception as e:
        log(f"Error updating answer_challenge.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_asset_ref_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 15: Update asset_ref.xlsx — clone asset blocks and achievement rows."""
    import openpyxl
    import stat
    import os
    import re
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "asset_ref.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Error: asset_ref.xlsx not found at {file_path}")
        return False
    
    # Create backup for undo
    backup_path = file_path + ".bak"
    import stat
    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
    shutil.copy2(file_path, backup_path)
    record_file_creation(log_dir, step_index, backup_path)
    
    # Name variants
    src_lower = src_evt.lower()
    src_upper = src_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:]
    tgt_lower = tgt_evt.lower()
    tgt_upper = tgt_evt.upper()
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:]
    src_digits = re.search(r'\d+', src_evt)
    tgt_digits = re.search(r'\d+', tgt_evt)
    src_digits = src_digits.group() if src_digits else ""
    tgt_digits = tgt_digits.group() if tgt_digits else ""
    
    def swap_name(val):
        if val is None:
            return val
        s = str(val)
        if isinstance(val, (int, float)):
            return val
        s = s.replace(src_upper, tgt_upper)
        s = s.replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal)
        s = s.replace(src_evt, tgt_evt)
        if src_digits and tgt_digits:
            s = s.replace(src_digits, tgt_digits)
        return s
    
    try:
        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        ws_main = wb["key_asset"] if "key_asset" in wb.sheetnames else None
        ws_backup = wb["key_asset备份"] if "key_asset备份" in wb.sheetnames else None
        
        if not ws_main:
            log("Error: key_asset sheet not found in asset_ref.xlsx")
            return False
        
        # === Part 1: Clone source event asset block ===
        def find_event_block(ws, event_upper):
            """Find all rows between //EVENT divider and the next divider or end."""
            block = []
            found_start = False
            for r in range(1, ws.max_row + 1):
                cell1 = str(ws.cell(r, 1).value or "")  # type: ignore
                normalized = cell1.strip().replace(" ", "").upper()
                if normalized == f"//{event_upper}" or normalized == f"//{event_upper}活动":
                    found_start = True
                    row_data = []
                    for c in range(1, ws.max_column + 1):
                        row_data.append(ws.cell(r, c).value)  # type: ignore
                    block.append(row_data)
                    continue
                if found_start:
                    if cell1.startswith("//"):
                        break
                    if cell1.strip() == "":
                        break
                    row_data = []
                    for c in range(1, ws.max_column + 1):
                        row_data.append(ws.cell(r, c).value)  # type: ignore
                    block.append(row_data)
            return block
        
        # Search key_asset first, then backup
        src_block_rows = find_event_block(ws_main, src_upper)
        block_source = "key_asset"
        
        if not src_block_rows and ws_backup:
            src_block_rows = find_event_block(ws_backup, src_upper)
            block_source = "key_asset备份"
        
        if src_block_rows:
            log(f"asset_ref [key_asset]: Found {len(src_block_rows)} row asset block for {src_evt} in {block_source}")
            
            # Check if target event block already exists
            existing_check = find_event_block(ws_main, tgt_upper)
            if existing_check:
                log(f"asset_ref [key_asset]: Target block //{tgt_upper} already exists. Skipping block clone.")
            else:
                # Find insertion point — before //活动成就资源 section
                insert_row = ws_main.max_row + 1
                for r in range(1, ws_main.max_row + 1):
                    cell1 = str(ws_main.cell(r, 1).value or "")  # type: ignore
                    if "活动成就" in cell1:
                        insert_row = r
                        break
                
                # Shift existing rows down to make space
                num_new_rows = len(src_block_rows)
                for r in range(ws_main.max_row, insert_row - 1, -1):
                    for c in range(1, ws_main.max_column + 1):
                        ws_main.cell(r + num_new_rows, c, ws_main.cell(r, c).value)  # type: ignore
                
                # Clear original cells to avoid duplicates
                for r in range(insert_row, insert_row + num_new_rows):
                    for c in range(1, ws_main.max_column + 1):
                        ws_main.cell(r, c, None)  # type: ignore
                
                # Insert cloned block
                for i, row_data in enumerate(src_block_rows):
                    target_r = insert_row + i
                    for c_idx, val in enumerate(row_data):
                        ws_main.cell(target_r, c_idx + 1, swap_name(val))  # type: ignore
                
                log(f"asset_ref [key_asset]: Inserted {num_new_rows} rows at R{insert_row} for //{tgt_upper}")
                modified = True
        else:
            log(f"asset_ref [key_asset]: No asset block found for {src_evt}. Normal for some event types.")
        
        # === Part 2: Clone achievement row ===
        achieve_src_row = None
        max_chronicle = 0
        
        for r in range(1, ws_main.max_row + 1):
            cell1 = str(ws_main.cell(r, 1).value or "")  # type: ignore
            if cell1.strip().upper() == f"ACHIEVEMENT_E_{src_upper}":
                achieve_src_row = r
            cell3 = str(ws_main.cell(r, 3).value or "")  # type: ignore
            chronicle_match = re.search(r'icon_Chronicle(\d+)', cell3, re.IGNORECASE)
            if chronicle_match:
                num = int(chronicle_match.group(1))
                if num > max_chronicle:
                    max_chronicle = num
        
        if achieve_src_row:
            # Check if target already exists
            tgt_achieve_exists = False
            for r in range(1, ws_main.max_row + 1):
                cell1 = str(ws_main.cell(r, 1).value or "")  # type: ignore
                if cell1.strip().upper() == f"ACHIEVEMENT_E_{tgt_upper}":
                    tgt_achieve_exists = True
                    break
            
            if tgt_achieve_exists:
                log(f"asset_ref [achievement]: ACHIEVEMENT_E_{tgt_upper} already exists. Skipping.")
            else:
                new_chronicle = max_chronicle + 1
                # Insert after the last achievement row (find //通用活动商店 section)
                insert_after = achieve_src_row
                for r in range(achieve_src_row + 1, ws_main.max_row + 1):
                    cell1 = str(ws_main.cell(r, 1).value or "")  # type: ignore
                    if cell1.startswith("ACHIEVEMENT_E_"):
                        insert_after = r
                    elif cell1.startswith("//"):
                        break
                
                new_row = insert_after + 1
                
                # Shift rows down
                for r in range(ws_main.max_row, new_row - 1, -1):
                    for c in range(1, ws_main.max_column + 1):
                        ws_main.cell(r + 1, c, ws_main.cell(r, c).value)  # type: ignore
                
                # Clone the source achievement row with name swap
                for c in range(1, ws_main.max_column + 1):
                    old_val = ws_main.cell(achieve_src_row, c).value  # type: ignore
                    new_val = swap_name(old_val)
                    if c == 3 and isinstance(new_val, str):
                        new_val = re.sub(r'icon_Chronicle\d+', f'icon_Chronicle{new_chronicle}', new_val)
                    ws_main.cell(new_row, c, new_val)  # type: ignore
                
                log(f"asset_ref [achievement]: Added ACHIEVEMENT_E_{tgt_upper} at R{new_row} with Chronicle{new_chronicle}")
                modified = True
        else:
            log(f"asset_ref [achievement]: No source achievement row found for {src_evt}")
        
        if modified:
            wb.save(file_path)
            log(f"asset_ref.xlsx successfully updated for {tgt_evt}.")
        else:
            log(f"asset_ref.xlsx: No changes needed.")
        
        return True
        
    except Exception as e:
        log(f"Error updating asset_ref.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_store_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 16: Update store.xlsx — clone event rows across 4 sheets with ID remapping."""
    import openpyxl
    import stat
    import os
    import re
    import shutil
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "store.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Error: store.xlsx not found at {file_path}")
        return False
    
    # Create backup for undo
    backup_path = file_path + ".bak"
    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
    shutil.copy2(file_path, backup_path)
    record_file_creation(log_dir, step_index, backup_path)
    
    # Name variants for swapping
    src_lower = src_evt.lower()
    src_upper = src_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:]
    tgt_lower = tgt_evt.lower()
    tgt_upper = tgt_evt.upper()
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:]
    
    def swap_name(val):
        """Replace source event name references with target event name."""
        if val is None:
            return val
        if isinstance(val, (int, float)):
            return val
        s = str(val)
        s = s.replace(src_upper, tgt_upper)
        s = s.replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal)
        s = s.replace(src_evt, tgt_evt)
        return s
    
    def remap_ids_in_cell(val, id_map):
        """Replace old IDs with new IDs in a comma-separated cell value."""
        if val is None:
            return val
        if isinstance(val, (int, float)):
            old_id = str(int(val))
            return int(id_map[old_id]) if old_id in id_map else val
        s = str(val)
        parts = s.split(",")
        new_parts = []
        for part in parts:
            stripped = part.strip()
            if stripped in id_map:
                new_parts.append(id_map[stripped])
            else:
                new_parts.append(part)
        return ",".join(new_parts)
    
    # Load extra_args for event timing
    extra_path = os.path.join(log_dir, "extra_args.json")
    start_time = None
    end_time = None
    if os.path.exists(extra_path):
        import json
        with open(extra_path, "r", encoding="utf-8") as f:
            extra = json.load(f)
        start_time = extra.get("start_time", "")
        end_time = extra.get("end_time", "")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        # =====================================================
        # Part 1: store_recharge — clone 4 event shop rows
        # =====================================================
        ws_r = wb["store_recharge"] if "store_recharge" in wb.sheetnames else None
        new_recharge_ids = []  # Track new IDs for store_package mirroring
        old_recharge_ids = []
        
        if ws_r:
            # Find source event rows
            src_recharge_rows = []
            for r in range(2, ws_r.max_row + 1):
                alias = str(ws_r.cell(r, 2).value or "")  # type: ignore
                if src_evt.upper() in alias.upper() or src_evt.lower() in alias.lower() or src_evt in alias:
                    # Only match event shop rows (商店礼包), not templates
                    if not alias.startswith("B-") and not alias.startswith("A-"):
                        src_recharge_rows.append(r)
                        old_recharge_ids.append(str(ws_r.cell(r, 1).value or ""))  # type: ignore
            
            if src_recharge_rows:
                # Check if target already exists
                tgt_exists = False
                for r in range(2, ws_r.max_row + 1):
                    alias = str(ws_r.cell(r, 2).value or "")  # type: ignore
                    if tgt_evt in alias and not alias.startswith("B-") and not alias.startswith("A-"):
                        tgt_exists = True
                        break
                
                if tgt_exists:
                    log(f"store_recharge: Target event {tgt_evt} already exists. Skipping.")
                else:
                    # Find max ID in same event section as source
                    # Source rows define the section; find next section divider after them
                    last_src_row = src_recharge_rows[-1]
                    section_end = last_src_row
                    for r in range(last_src_row + 1, ws_r.max_row + 1):
                        id_val = str(ws_r.cell(r, 1).value or "")  # type: ignore
                        if id_val.startswith("//"):
                            section_end = r - 1
                            break
                        # Also stop if we hit an empty row
                        if not id_val:
                            section_end = r - 1
                            break
                        section_end = r
                    
                    # Find max ID within the event section (between // markers)
                    # Walk backward from source to find section start marker
                    section_start = 2
                    for r in range(src_recharge_rows[0] - 1, 1, -1):
                        id_val = str(ws_r.cell(r, 1).value or "")  # type: ignore
                        if id_val.startswith("//"):
                            section_start = r + 1
                            break
                    
                    max_id = 0
                    for r in range(section_start, section_end + 1):
                        id_val = ws_r.cell(r, 1).value  # type: ignore
                        if id_val:
                            try:
                                num = int(id_val)
                                if num > max_id:
                                    max_id = num
                            except (ValueError, TypeError):
                                pass
                    
                    # Insert right after last data row in the event section
                    insert_row = section_end + 1
                    
                    next_id = max_id + 1
                    for i, src_r in enumerate(src_recharge_rows):
                        new_id = next_id + i
                        new_recharge_ids.append(str(new_id))
                        for c in range(1, ws_r.max_column + 1):
                            old_val = ws_r.cell(src_r, c).value  # type: ignore
                            new_val = swap_name(old_val)
                            if c == 1:
                                new_val = new_id
                            ws_r.cell(insert_row + i, c, new_val)  # type: ignore
                    
                    log(f"store_recharge: Cloned {len(src_recharge_rows)} rows at R{insert_row}. IDs {new_recharge_ids[0]}-{new_recharge_ids[-1]}")
                    modified = True
            else:
                log(f"store_recharge: No event rows found for {src_evt}")
        
        # =====================================================
        # Part 2: store_package — clone with SAME IDs as recharge
        # =====================================================
        ws_p = wb["store_package"] if "store_package" in wb.sheetnames else None
        
        if ws_p and new_recharge_ids:
            src_package_rows = []
            for r in range(2, ws_p.max_row + 1):
                id_val = str(ws_p.cell(r, 1).value or "")  # type: ignore
                if id_val in old_recharge_ids:
                    src_package_rows.append(r)
            
            if src_package_rows:
                # Check if target already exists
                tgt_exists = any(
                    str(ws_p.cell(r, 1).value or "") in new_recharge_ids  # type: ignore
                    for r in range(2, ws_p.max_row + 1)
                )
                
                if tgt_exists:
                    log(f"store_package: Target IDs already exist. Skipping.")
                else:
                    # Insert after last source event row in the section
                    last_pkg_src = src_package_rows[-1]
                    pkg_section_end = last_pkg_src
                    for r in range(last_pkg_src + 1, ws_p.max_row + 1):
                        id_val = str(ws_p.cell(r, 1).value or "")  # type: ignore
                        if id_val.startswith("//") or not id_val:
                            pkg_section_end = r - 1
                            break
                        pkg_section_end = r
                    
                    insert_row = pkg_section_end + 1
                    
                    old_to_new_recharge = dict(zip(old_recharge_ids, new_recharge_ids))
                    
                    for i, src_r in enumerate(src_package_rows):
                        for c in range(1, ws_p.max_column + 1):
                            old_val = ws_p.cell(src_r, c).value  # type: ignore
                            new_val = swap_name(old_val)
                            if c == 1:
                                old_id = str(old_val)
                                new_val = int(old_to_new_recharge.get(old_id, old_val))
                            ws_p.cell(insert_row + i, c, new_val)  # type: ignore
                    
                    log(f"store_package: Cloned {len(src_package_rows)} rows at R{insert_row} with matching IDs")
                    modified = True
            else:
                log(f"store_package: No rows found matching recharge IDs for {src_evt}")
        
        # =====================================================
        # Part 3: recommend_gift — clone 18 rows with ID remapping
        # =====================================================
        ws_g = wb["recommend_gift"] if "recommend_gift" in wb.sheetnames else None
        new_gift_ids = []
        old_gift_ids = []
        gift_id_map = {}  # old_id → new_id for internal remapping
        
        if ws_g:
            src_gift_rows = []
            for r in range(2, ws_g.max_row + 1):
                row_text = ""
                for c in [2, 21, 29]:  # Alias, 依赖活动, 推荐条件
                    row_text += " " + str(ws_g.cell(r, c).value or "")  # type: ignore
                if src_evt.upper() in row_text.upper() or src_evt in row_text:
                    id_val = ws_g.cell(r, 1).value  # type: ignore
                    if id_val and not str(id_val).startswith("//"):
                        src_gift_rows.append(r)
                        old_gift_ids.append(str(int(id_val)))
            
            if src_gift_rows:
                # Check target exists
                tgt_exists = False
                for r in range(2, ws_g.max_row + 1):
                    alias = str(ws_g.cell(r, 2).value or "")  # type: ignore
                    if tgt_evt in alias:
                        tgt_exists = True
                        break
                
                if tgt_exists:
                    log(f"recommend_gift: Target event {tgt_evt} already exists. Skipping.")
                else:
                    # Find max ID only in the event gift section (row 61+)
                    # The event section starts after "// 活动商店与进度礼包" row
                    # Event gift IDs are in the 9000-12999 range
                    # Non-event IDs (18xxx for 周末礼包, 芭比 etc.) are in rows 16-60
                    event_section_start = 61
                    for r in range(2, ws_g.max_row + 1):
                        id_val = str(ws_g.cell(r, 1).value or "")  # type: ignore
                        if id_val.startswith("// 活动商店"):
                            event_section_start = r + 1
                            break
                    
                    max_gift_id = 0
                    for r in range(event_section_start, ws_g.max_row + 1):
                        id_val = ws_g.cell(r, 1).value  # type: ignore
                        if id_val:
                            try:
                                num = int(id_val)
                                if num > max_gift_id:
                                    max_gift_id = num
                            except (ValueError, TypeError):
                                pass
                    
                    log(f"recommend_gift: Event section max ID = {max_gift_id} (from R{event_section_start}+)")
                    
                    # Build old→new ID map
                    for i, old_id in enumerate(old_gift_ids):
                        new_id = str(max_gift_id + 1 + i)
                        gift_id_map[old_id] = new_id
                        new_gift_ids.append(new_id)
                    
                    log(f"recommend_gift: ID map ({len(gift_id_map)} entries): {list(gift_id_map.items())[:4]}...")
                    
                    # Insert after the last event row (not at sheet end)
                    last_gift_src = src_gift_rows[-1]
                    insert_row = last_gift_src + 1
                    # Walk forward to find the actual last event row
                    for r in range(ws_g.max_row, last_gift_src, -1):
                        alias = str(ws_g.cell(r, 2).value or "")  # type: ignore
                        id_val = ws_g.cell(r, 1).value  # type: ignore
                        if id_val and not str(id_val).startswith("//") and alias:
                            insert_row = r + 1
                            break
                    
                    # Derive end_time from source event's actual end_time (23:44:59 pattern)
                    src_end_time = None
                    src_start_time = None
                    if src_gift_rows:
                        src_end_time = ws_g.cell(src_gift_rows[0], 13).value  # type: ignore
                        src_start_time = ws_g.cell(src_gift_rows[0], 12).value  # type: ignore
                    
                    # Columns that need internal ID remapping
                    remap_cols = {4, 7, 16}  # 合并购买关系, 互斥组, 属性继承
                    
                    for i, src_r in enumerate(src_gift_rows):
                        for c in range(1, ws_g.max_column + 1):
                            old_val = ws_g.cell(src_r, c).value  # type: ignore
                            
                            if c == 1:
                                # Replace ID
                                old_id_str = str(int(old_val))
                                new_val = int(gift_id_map.get(old_id_str, old_val))
                            elif c in remap_cols:
                                # Remap internal cross-references
                                new_val = remap_ids_in_cell(old_val, gift_id_map)
                            elif c == 12 and start_time:
                                # Start time from extra_args
                                new_val = start_time
                            elif c == 13 and src_end_time is not None:
                                # Derive end_time: use target event end date + source time portion
                                # Source end_time has the correct time portion (e.g. 23:44:59)
                                src_end_str = str(src_end_time)
                                if end_time and " " in src_end_str:
                                    # Use end date from extra_args + time from source
                                    src_time_part = src_end_str.split(" ")[-1]
                                    end_date = str(end_time).split(" ")[0] if " " in str(end_time) else str(end_time)
                                    new_val = f"{end_date} {src_time_part}"
                                else:
                                    new_val = swap_name(old_val)
                            else:
                                new_val = swap_name(old_val)
                            
                            ws_g.cell(insert_row + i, c, new_val)  # type: ignore
                    
                    log(f"recommend_gift: Cloned {len(src_gift_rows)} rows at R{insert_row}. IDs {new_gift_ids[0]}-{new_gift_ids[-1]}")
                    modified = True
            else:
                log(f"recommend_gift: No rows found for {src_evt}")
        
        # =====================================================
        # Part 4: event_gift_style — clone 12 rows (first 12 gift IDs)
        # =====================================================
        ws_s = wb["event_gift_style"] if "event_gift_style" in wb.sheetnames else None
        
        if ws_s and gift_id_map:
            # Only first 12 of 18 recommend_gift IDs have style entries
            main_old_ids = old_gift_ids[:12] if len(old_gift_ids) >= 12 else old_gift_ids
            
            src_style_rows = []
            for r in range(2, ws_s.max_row + 1):
                id_val = str(ws_s.cell(r, 1).value or "")  # type: ignore
                if id_val in main_old_ids:
                    src_style_rows.append(r)
            
            if src_style_rows:
                # Check target exists
                main_new_ids = [gift_id_map[oid] for oid in main_old_ids if oid in gift_id_map]
                tgt_exists = any(
                    str(ws_s.cell(r, 1).value or "") in main_new_ids  # type: ignore
                    for r in range(2, ws_s.max_row + 1)
                )
                
                if tgt_exists:
                    log(f"event_gift_style: Target IDs already exist. Skipping.")
                else:
                    insert_row = ws_s.max_row + 1
                    for r in range(ws_s.max_row, 1, -1):
                        if ws_s.cell(r, 1).value is not None:  # type: ignore
                            insert_row = r + 1
                            break
                    
                    for i, src_r in enumerate(src_style_rows):
                        for c in range(1, ws_s.max_column + 1):
                            old_val = ws_s.cell(src_r, c).value  # type: ignore
                            if c == 1:
                                old_id_str = str(old_val)
                                new_val = int(gift_id_map.get(old_id_str, old_val))
                            else:
                                new_val = old_val  # Style values don't contain event names
                            ws_s.cell(insert_row + i, c, new_val)  # type: ignore
                    
                    log(f"event_gift_style: Cloned {len(src_style_rows)} rows")
                    modified = True
            else:
                log(f"event_gift_style: No style rows found for source gift IDs")
        
        if modified:
            wb.save(file_path)
            log(f"store.xlsx successfully updated for {tgt_evt}.")
        else:
            log(f"store.xlsx: No changes needed.")
        
        return True
        
    except Exception as e:
        log(f"Error updating store.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_pack_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 18: Update pack.xlsx — clone pack reward rows and auto-increment IDs."""
    import openpyxl
    import stat
    import os
    import re
    import shutil
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "pack.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Error: pack.xlsx not found at {file_path}")
        return False
        
    backup_path = file_path + ".bak"
    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
    shutil.copy2(file_path, backup_path)
    record_file_creation(log_dir, step_index, backup_path)
    
    src_lower = src_evt.lower()
    src_upper = src_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:]
    tgt_lower = tgt_evt.lower()
    tgt_upper = tgt_evt.upper()
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:]
    
    def swap_name(val):
        """Replace source event name references with target event name."""
        if val is None:
            return val
        if isinstance(val, (int, float)):
            return val
        s = str(val)
        s = s.replace(src_upper, tgt_upper)
        s = s.replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal)
        s = s.replace(src_evt, tgt_evt)
        return s

    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        ws = wb["pack"] if "pack" in wb.sheetnames else None
        if not ws:
            log(f"pack.xlsx: 'pack' sheet not found")
            return False
            
        src_rows = []
        for r in range(2, ws.max_row + 1):
            alias = str(ws.cell(r, 2).value or "")  # type: ignore
            if src_evt.upper() in alias.upper() or src_evt in alias:
                src_rows.append(r)
                
        if not src_rows:
            log(f"pack.xlsx: No source rows found for {src_evt}")
            return True
            
        tgt_exists = False
        for r in range(2, ws.max_row + 1):
            alias = str(ws.cell(r, 2).value or "")  # type: ignore
            if tgt_evt in alias:
                tgt_exists = True
                break
                
        if tgt_exists:
            log(f"pack.xlsx: Target event {tgt_evt} already exists. Skipping.")
            return True
            
        max_id = 0
        for r in range(2, ws.max_row + 1):
            id_val = ws.cell(r, 1).value  # type: ignore
            if id_val:
                try:
                    num = int(id_val)
                    if num > max_id:
                        max_id = num
                except (ValueError, TypeError):
                    pass
                    
        insert_row = ws.max_row + 1
        for r in range(ws.max_row, 1, -1):
            if ws.cell(r, 1).value is not None:  # type: ignore
                insert_row = r + 1
                break
                
        next_id = max_id + 1
        for i, src_r in enumerate(src_rows):
            new_id = next_id + i
            for c in range(1, ws.max_column + 1):
                old_val = ws.cell(src_r, c).value  # type: ignore
                new_val = swap_name(old_val)
                if c == 1:
                    new_val = new_id
                ws.cell(insert_row + i, c, new_val)  # type: ignore
                
        log(f"pack.xlsx: Cloned {len(src_rows)} rows. New IDs {next_id} to {next_id + len(src_rows) - 1}")
        
        wb.save(file_path)
        log(f"pack.xlsx successfully updated for {tgt_evt}.")
        return True
        
    except Exception as e:
        log(f"Error updating pack.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_guide_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 19: Update guide.xlsx — clone, remap IDs, insert precisely after active events."""
    import openpyxl
    import stat
    import os
    import shutil
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    file_path = os.path.join(design_dir, "guide.xlsx")
    
    if not os.path.exists(file_path):
        log(f"Error: guide.xlsx not found at {file_path}")
        return False
        
    backup_path = file_path + ".bak"
    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
    shutil.copy2(file_path, backup_path)
    record_file_creation(log_dir, step_index, backup_path)
    
    src_lower = src_evt.lower()
    src_upper = src_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:]
    tgt_lower = tgt_evt.lower()
    tgt_upper = tgt_evt.upper()
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:]
    
    def swap_name(val):
        if val is None: return val
        if isinstance(val, (int, float)): return val
        s = str(val)
        s = s.replace(src_upper, tgt_upper)
        s = s.replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal)
        s = s.replace(src_evt, tgt_evt)
        return s

    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        # 1. guide_section
        old_id_to_new_id = {}
        ws_sec = wb["guide_section"] if "guide_section" in wb.sheetnames else None
        if ws_sec:
            # find max ID among the event sections (IDs < 100000)
            max_evt_id = 0
            insert_sec_row = ws_sec.max_row + 1
            start_commenting_idx = 2
            
            for r in range(2, ws_sec.max_row + 1):
                id_val_raw = str(ws_sec.cell(r, 1).value or "")
                id_val = id_val_raw.replace("//", "")
                
                if id_val == "619":
                    start_commenting_idx = r
                    
                if id_val.isdigit():
                    num = int(id_val)
                    if num < 100000:
                        max_evt_id = max(max_evt_id, num)
                    elif num >= 100000 and insert_sec_row > ws_sec.max_row:
                        # Found the start of the core guides (100001)
                        # We will insert standard events right before this row, 
                        # but wait, let's insert right before the //新手固定活动 divider if it exists
                        # Let's peek backwards to see if there's a divider
                        insert_sec_row = r
                        if str(ws_sec.cell(r - 1, 1).value or "").startswith("//"):
                            insert_sec_row = r - 1
            
            src_rows = []
            for r in range(2, ws_sec.max_row + 1):
                alias = str(ws_sec.cell(r, 2).value or "")
                id_val = str(ws_sec.cell(r, 1).value or "")
                if src_evt.upper() in alias.upper() or src_evt in alias:
                    src_rows.append(r)
            
            if src_rows:
                next_id = max_evt_id + 1
                for i, src_r in enumerate(src_rows):
                    old_id_str = str(ws_sec.cell(src_r, 1).value or "").replace("//", "")
                    if old_id_str.isdigit():
                        old_id_to_new_id[int(old_id_str)] = next_id + i
                        
                # Insert rows precisely before the core game guides (insert_sec_row)
                ws_sec.insert_rows(insert_sec_row, amount=len(src_rows))
                
                # Fill newly inserted rows
                for i, src_r in enumerate(src_rows):
                    actual_src_r = src_r if src_r < insert_sec_row else src_r + len(src_rows)
                    for c in range(1, ws_sec.max_column + 1):
                        old_val = ws_sec.cell(actual_src_r, c).value
                        new_val = swap_name(old_val)
                        if c == 1 and new_val:
                            old_id_str = str(old_val).replace("//", "")
                            if old_id_str.isdigit() and int(old_id_str) in old_id_to_new_id:
                                parsed_id = int(old_id_to_new_id[int(old_id_str)])
                                new_val = str(parsed_id) if isinstance(old_val, str) else parsed_id
                        ws_sec.cell(insert_sec_row + i, c, new_val)
                        
                # Comment out previous events starting from 619 up to our new insertion
                for r in range(start_commenting_idx, insert_sec_row):
                    val = ws_sec.cell(r, 1).value
                    if val and not str(val).startswith("//"):
                        ws_sec.cell(r, 1, f"//{val}")
                
                modified = True
                log(f"guide_section: Inserted at row {insert_sec_row}. Mapped {len(old_id_to_new_id)} IDs.")
        
        # 2. guide and guide_backup
        ws_guide = wb["guide"] if "guide" in wb.sheetnames else None
        ws_gb = wb["guide_backup"] if "guide_backup" in wb.sheetnames else None
        
        if ws_guide and ws_gb and old_id_to_new_id:
            # Guide event IDs are sectionID * 100 + step. e.g. section 619 -> guide ID > 61900
            # Identify active event rows in guide (61900 <= ID < 10000000)
            rows_to_transfer = []
            insert_g_row = ws_guide.max_row + 1
            
            for r in range(2, ws_guide.max_row + 1):
                id_val_str = str(ws_guide.cell(r, 1).value or "").replace("//", "")
                if id_val_str.isdigit():
                    num = int(id_val_str)
                    if 61900 <= num < 10000000:
                        rows_to_transfer.append(r)
                    elif num >= 10000000 and insert_g_row > ws_guide.max_row:
                        insert_g_row = r
                        if str(ws_guide.cell(r - 1, 1).value or "").startswith("//"):
                            insert_g_row = r - 1
            
            # If no rows to transfer were found, we still need the insertion point
            if insert_g_row > ws_guide.max_row:
                for r in range(2, ws_guide.max_row + 1):
                    id_val_str = str(ws_guide.cell(r, 1).value or "").replace("//", "")
                    if id_val_str.isdigit() and int(id_val_str) >= 10000000:
                        insert_g_row = r
                        if str(ws_guide.cell(r - 1, 1).value or "").startswith("//"):
                            insert_g_row = r - 1
                        break
            
            if rows_to_transfer:
                gb_insert_row = ws_gb.max_row + 1
                for r in range(ws_gb.max_row, 1, -1):
                    if ws_gb.cell(r, 1).value is not None:
                        gb_insert_row = r + 1
                        break
                        
                for r in rows_to_transfer:
                    id_val = str(ws_guide.cell(r, 1).value or "")
                    exists_in_gb = False
                    # Simple check to avoid duplicates in backup
                    for b_row in range(gb_insert_row - 100, gb_insert_row):
                        if b_row > 1 and str(ws_gb.cell(b_row, 1).value or "") == id_val:
                            exists_in_gb = True
                            break
                    if not exists_in_gb and id_val:
                        for c in range(1, ws_guide.max_column + 1):
                            ws_gb.cell(gb_insert_row, c, ws_guide.cell(r, c).value)
                        gb_insert_row += 1
                
                # Delete rows from guide
                # Do it in reverse order to preserve indices
                for r in reversed(rows_to_transfer):
                    ws_guide.delete_rows(r, 1)
                    if r < insert_g_row:
                        insert_g_row -= 1
            
            # Find source rows in guide_backup using old_id_to_new_id prefix matching
            src_gb_rows = []
            for r in range(2, ws_gb.max_row + 1):
                id_val_str = str(ws_gb.cell(r, 1).value or "").replace("//", "")
                if id_val_str.isdigit() and len(id_val_str) >= 3:
                    base_old = int(id_val_str[:-2])
                    if base_old in old_id_to_new_id:
                        src_gb_rows.append(r)
            
            if src_gb_rows:
                ws_guide.insert_rows(insert_g_row, amount=len(src_gb_rows))
                for i, src_r in enumerate(src_gb_rows):
                    for c in range(1, ws_gb.max_column + 1):
                        old_val = ws_gb.cell(src_r, c).value
                        new_val = swap_name(old_val)
                        if c == 1 and new_val:
                            old_id_str = str(old_val).replace("//", "")
                            if old_id_str.isdigit() and len(old_id_str) >= 3:
                                base_old = int(old_id_str[:-2])
                                remng = old_id_str[-2:]
                                if base_old in old_id_to_new_id:
                                    mapped = int(f"{old_id_to_new_id[base_old]}{remng}")
                                    new_val = str(mapped) if isinstance(old_val, str) else mapped
                        ws_guide.cell(insert_g_row + i, c, new_val)
                modified = True
                log(f"guide & guide_backup: Transferred {len(rows_to_transfer)} rows, Cloned {len(src_gb_rows)} rows to row {insert_g_row}.")
                        
        # 3. guide_trigger
        ws_trig = wb["guide_trigger"] if "guide_trigger" in wb.sheetnames else None
        if ws_trig:
            # find original rows by name matching 
            src_rows = []
            for r in range(2, ws_trig.max_row + 1):
                alias = str(ws_trig.cell(r, 2).value or "")
                id_val = str(ws_trig.cell(r, 1).value or "")
                if src_evt.upper() in alias.upper() or src_evt in alias or src_evt.upper() in id_val.upper() or src_evt in id_val:
                    src_rows.append(r)
                    
            if src_rows:
                insert_row = ws_trig.max_row + 1
                for r in range(ws_trig.max_row, 1, -1):
                    if ws_trig.cell(r, 1).value is not None:
                        insert_row = r + 1
                        break
                
                # We append to the bottom for triggers 
                for i, src_r in enumerate(src_rows):
                    for c in range(1, ws_trig.max_column + 1):
                        old_val = ws_trig.cell(src_r, c).value
                        new_val = swap_name(old_val)
                        if c == 1 and new_val:
                            new_val = str(new_val).replace("//", "") # Uncomment
                        ws_trig.cell(insert_row + i, c, new_val)
                        
                start_commenting_row = 148
                for r in range(start_commenting_row, insert_row):
                    val = ws_trig.cell(r, 1).value
                    if val and not str(val).startswith("//"):
                        s_val = str(val).upper()
                        # DO NOT comment out shared MergeFix triggers or global triggers
                        if "MERGE_FIX" in s_val or "MERGEFIX" in s_val:
                            continue
                        ws_trig.cell(r, 1, f"//{val}")
                modified = True
                        
        # 4. guide_refui
        ws_ui = wb["guide_refui"] if "guide_refui" in wb.sheetnames else None
        if ws_ui:
            src_rows = []
            for r in range(2, ws_ui.max_row + 1):
                alias = str(ws_ui.cell(r, 2).value or "")
                id_val = str(ws_ui.cell(r, 1).value or "")
                if src_evt.upper() in alias.upper() or src_evt in alias or src_evt.upper() in id_val.upper() or src_evt in id_val:
                    src_rows.append(r)
                    
            if src_rows:
                insert_row = ws_ui.max_row + 1
                for r in range(ws_ui.max_row, 1, -1):
                    if ws_ui.cell(r, 1).value is not None:
                        insert_row = r + 1
                        break
                        
                for i, src_r in enumerate(src_rows):
                    for c in range(1, ws_ui.max_column + 1):
                        old_val = ws_ui.cell(src_r, c).value
                        new_val = swap_name(old_val)
                        if c == 1 and new_val:
                            new_val = str(new_val).replace("//", "") # Uncomment
                        ws_ui.cell(insert_row + i, c, new_val)
                        
                start_commenting_row = 97
                for r in range(start_commenting_row, insert_row):
                    val = ws_ui.cell(r, 1).value
                    if val and not str(val).startswith("//"):
                        s_val = str(val).upper()
                        # DO NOT comment out shared MergeFix UI or global UI refs
                        if "MERGE_FIX" in s_val or "MERGEFIX" in s_val:
                            continue
                        ws_ui.cell(r, 1, f"//{val}")
                modified = True

        if modified:
            wb.save(file_path)
            log(f"guide.xlsx successfully updated for {tgt_evt}.")
        else:
            log(f"guide.xlsx: No changes needed.")
            
        return True
    
    except Exception as e:
        log(f"Error updating guide.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_localization_all(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 14: Orchestrator for all localization and quiz updates."""
    
    # Part A: localization.xlsx
    log("=== Step 14 Part A: localization.xlsx ===")
    success_a = update_localization_main_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
    if not success_a:
        return False
    
    # Part B: localization_quiz.xlsx
    log("=== Step 14 Part B: localization_quiz.xlsx ===")
    success_b = update_localization_quiz_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
    if not success_b:
        return False
    
    # Part C: answer_challenge.xlsx
    log("=== Step 14 Part C: answer_challenge.xlsx ===")
    success_c = update_answer_challenge_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
    if not success_c:
        return False
    
    log("=== Step 14 Complete: All localization and quiz files updated. ===")
    return True

def update_sys_excel(proj_root, src_evt, tgt_evt, log_dir, step_index):
    """Step 19.5/20: Update sys.xlsx — register DeepLink IDs in correct sections."""
    import openpyxl
    import stat
    import os
    import shutil
    
    design_dir = os.path.join(proj_root, r"design\DesignData\WorkSpcae_Design\design")
    sys_path = os.path.join(design_dir, "sys.xlsx")
    events_path = os.path.join(design_dir, "events.xlsx")
    tgt_evt_xlsx = os.path.join(design_dir, f"{tgt_evt}.xlsx")
    
    if not os.path.exists(sys_path):
        log(f"Error: sys.xlsx not found at {sys_path}")
        return False
        
    # 1. Fetch the New Achievement DeepLink ID from events.xlsx
    tgt_achievement_dl_id = None
    try:
        wb_evt = openpyxl.load_workbook(events_path, data_only=True)
        ws_a = wb_evt["活动成就"]
        for r in range(2, ws_a.max_row + 1):
            alias_val = str(ws_a.cell(r, 4).value or "").lower()
            if tgt_evt.lower() in alias_val:
                tgt_achievement_dl_id = ws_a.cell(r, 6).value
                break
        wb_evt.close()
    except Exception as e:
        log(f"sys.xlsx: Failed to read Achievement ID from events.xlsx: {e}")
        return False
        
    log(f"sys.xlsx: Target Achievement DeepLink ID: {tgt_achievement_dl_id}")

    # 2. Start Processing sys.xlsx
    backup_path = sys_path + ".bak"
    os.chmod(sys_path, stat.S_IWRITE | stat.S_IREAD)
    shutil.copy2(sys_path, backup_path)
    record_file_creation(log_dir, step_index, backup_path)
    
    src_lower, tgt_lower = src_evt.lower(), tgt_evt.lower()
    src_upper, tgt_upper = src_evt.upper(), tgt_evt.upper()
    src_pascal = src_evt[0].upper() + src_evt[1:] if src_evt else src_evt
    tgt_pascal = tgt_evt[0].upper() + tgt_evt[1:] if tgt_evt else tgt_evt

    def swap_name(val):
        if val is None or isinstance(val, (int, float)): return val
        s = str(val)
        s = s.replace(src_upper, tgt_upper).replace(src_lower, tgt_lower)
        s = s.replace(src_pascal, tgt_pascal).replace(src_evt, tgt_evt)
        return s

    try:
        # Step 2a: Scan for Max IDs and Dividers using data_only=True
        wb_scan = openpyxl.load_workbook(sys_path, data_only=True)
        ws_scan = wb_scan["DeepLink"]
        
        row_achieve_divider = None
        row_invite_divider = None
        max_share_id = 0
        src_share_rows = []
        src_achieve_row_idx = None

        for r in range(1, ws_scan.max_row + 1):
            val1 = ws_scan.cell(r, 1).value
            val2 = str(ws_scan.cell(r, 2).value or "") or ""
            
            # Robust ID extraction (handles both int and string IDs)
            id_num = None
            try:
                if val1 is not None:
                    id_num = int(float(str(val1)))
            except:
                pass
            
            # Normalize divider check
            s_val1 = str(val1).replace(" ", "") if val1 else ""
            if s_val1 == "//活动成就": row_achieve_divider = r
            if s_val1 == "//邀新": row_invite_divider = r
            
            # Find Source Rows
            if src_evt.lower() in val2.lower():
                if id_num is not None:
                    if id_num < 20000: src_share_rows.append(r)
                    else: src_achieve_row_idx = r
            
            # Find Section Max IDs (only before the first divider)
            if id_num is not None:
                if row_achieve_divider is None and id_num < 20000:
                    max_share_id = max(max_share_id, id_num)
                    
        wb_scan.close()

        log(f"sys.xlsx Sections: ShareDivider={row_achieve_divider}, InviteDivider={row_invite_divider}")
        log(f"sys.xlsx MaxIDs: Share={max_share_id}")

        if not row_achieve_divider:
            log("Error: Could not find '// 活动成就' divider in sys.xlsx")
            return False

        # Step 2b: Open for writing
        wb = openpyxl.load_workbook(sys_path)
        ws = wb["DeepLink"]

        # 3. Update Sim2605.xlsx if we have Share IDs to map
        if src_share_rows and os.path.exists(tgt_evt_xlsx):
            # Get the exact numeric values from scan result
            wb_ref = openpyxl.load_workbook(sys_path, data_only=True)
            ws_ref = wb_ref["DeepLink"]
            src_first_id = int(float(str(ws_ref.cell(src_share_rows[0], 1).value)))
            src_ids = [int(float(str(ws_ref.cell(r, 1).value))) for r in src_share_rows]
            wb_ref.close()

            new_first_id = max_share_id + 1
            offset = new_first_id - src_first_id
            
            # Use string versions for broad mapping in the event spreadsheet
            src_ids_str = [str(x) for x in src_ids]
            
            log(f"Mapping Share IDs: {src_ids[0]}..{src_ids[-1]} -> {new_first_id}.. (Offset: {offset})")
            
            wb_tgt = openpyxl.load_workbook(tgt_evt_xlsx)
            replaced_count = 0
            for sheet in wb_tgt.worksheets:
                for r in range(1, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        v = sheet.cell(r, c).value
                        if v is not None:
                            try:
                                v_int = int(float(str(v)))
                                if v_int in src_ids:
                                    sheet.cell(r, c, v_int + offset)
                                    replaced_count += 1
                            except:
                                pass
            wb_tgt.save(tgt_evt_xlsx)
            log(f"Updated {tgt_evt}.xlsx: {replaced_count} IDs remapped.")

        # 4. Insert new rows into sys.xlsx
        # Part B: Achievement ID (before // 邀新)
        if row_invite_divider and src_achieve_row_idx:
            # Check if group already exists (normalize type for check)
            if not any(str(ws.cell(r, 1).value).replace(" ","") == str(tgt_achievement_dl_id) for r in range(row_achieve_divider, row_invite_divider + 1)):
                ws.insert_rows(row_invite_divider)
                for c in range(1, ws.max_column + 1):
                    old_val = ws.cell(src_achieve_row_idx, c).value
                    new_val = swap_name(old_val)
                    if c == 1: new_val = int(tgt_achievement_dl_id)
                    ws.cell(row_invite_divider, c, new_val)
                log(f"Inserted Achievement ID {tgt_achievement_dl_id} at row {row_invite_divider}")

        # Part A: Share Group (before // 活动成就)
        if row_achieve_divider and src_share_rows:
            new_first_id = max_share_id + 1
            # Check if group already exists
            if not any(str(ws.cell(r, 1).value) == str(new_first_id) for r in range(1, row_achieve_divider + 1)):
                wb_ref = openpyxl.load_workbook(sys_path, data_only=True)
                ws_ref = wb_ref["DeepLink"]
                src_ref_ids = [int(float(str(ws_ref.cell(r, 1).value))) for r in src_share_rows]
                offset = new_first_id - src_ref_ids[0]
                wb_ref.close()

                for i, r_idx in enumerate(src_share_rows):
                    new_row_pos = row_achieve_divider + i
                    ws.insert_rows(new_row_pos)
                    for c in range(1, ws.max_column + 1):
                        old_val = ws.cell(r_idx, c).value
                        new_val = swap_name(old_val)
                        if c == 1: new_val = src_ref_ids[i] + offset
                        ws.cell(new_row_pos, c, new_val)
                log(f"Inserted {len(src_share_rows)} Share IDs starting at {new_first_id}")

        wb.save(sys_path)
        log("sys.xlsx successfully updated with sections and remapped IDs.")
        return True
    
    except Exception as e:
        log(f"Error updating sys.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    except Exception as e:
        log(f"Error updating sys.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    if len(sys.argv) < 6:
        log("Usage: python event_pipeline.py <step_index> <source_event> <target_event> <project_root> <is_revert>")
        sys.exit(1)

    step_index = int(sys.argv[1])
    src_evt = sys.argv[2]
    tgt_evt = sys.argv[3]
    proj_root = sys.argv[4]
    
    # Optional 5th arg indicating revert mode
    is_revert = True if sys.argv[5] == "1" else False

    log_dir = os.path.join(proj_root, r"develop\client\Skipbo\Assets\Editor\EventAutomation\PythonBackend")
    os.makedirs(log_dir, exist_ok=True)

    if is_revert:
        log(f"Reverting Python Step {step_index} | {src_evt} -> {tgt_evt}")
        success = revert_step(log_dir, step_index)
        sys.exit(0 if success else 1)
        
    log(f"Executing Python Step {step_index} | {src_evt} -> {tgt_evt}")

    if step_index == 0:
        proto_dir = os.path.join(proj_root, r"develop\protocol")
        success = inject_proto(proto_dir, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 2:
        # Step 3: Add active save file in local service
        src_file = os.path.join(proj_root, rf"develop\client\Skipbo\Assets\Lua\Game\Module\LocalService\Event\E{src_evt.capitalize()}LocalService.lua")
        tgt_file = os.path.join(proj_root, rf"develop\client\Skipbo\Assets\Lua\Game\Module\LocalService\Event\E{tgt_evt.capitalize()}LocalService.lua")
        success = copy_and_rename_file(src_file, tgt_file, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 3:
        # Step 4: Register in LocalServiceMgr.lua
        success = inject_local_service_mgr(proj_root, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 4:
        # Step 5: Clone WorkSpace3 configuration to Design
        success = clone_workspace3_to_design(proj_root, step_index)
        if success:
            design_dir = os.path.join(proj_root, r"design\DesignData")
            workspace_design = os.path.join(design_dir, "WorkSpcae_Design")
            record_file_creation(log_dir, step_index, workspace_design)
        if not success: sys.exit(1)
    elif step_index == 5:
        # Step 6: Clone Event Excel configuration tables
        success = clone_event_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 6:
        # Step 7: Copy event descriptor JSON(s) to app_client
        success = clone_event_descriptors(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 7:
        # Step 8: Modify convert_layout.lua and convert.json
        success = inject_convert_references(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 8:
        # Step 9: Update bi.xlsx
        success = update_bi_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 9:
        # Step 10: Update events.xlsx
        success = update_events_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 10:
        # Step 11: Update event_shop.xlsx
        success = update_event_shop_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 11:
        # Step 12: Update item.xlsx
        success, id_map = update_item_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if success and id_map:
            sync_events_excel_item_ids(proj_root, tgt_evt, id_map, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 12:
        # Step 13: Update icon.xlsx
        success = update_icon_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 13:
        # Step 14: Update localization files, quiz, and answer_challenge
        success = update_localization_all(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 14:
        # Step 15: Update asset_ref.xlsx
        success = update_asset_ref_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 15:
        # Step 16: Update store.xlsx
        success = update_store_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 16:
        # Step 17: Update card.xlsx (Placeholder)
        log("card.xlsx automation skipped (no event info present).")
    elif step_index == 17:
        # Step 18: Update pack.xlsx
        success = update_pack_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 18:
        # Step 19: Update guide.xlsx
        success = update_guide_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    elif step_index == 19:
        # Step 20: Update sys.xlsx
        success = update_sys_excel(proj_root, src_evt, tgt_evt, log_dir, step_index)
        if not success: sys.exit(1)
    else:
        log(f"Step {step_index} execution mock succeeded.")
        # Optional mock file creation to test undoing
        # mock_file = os.path.join(log_dir, f"mock_file_step_{step_index}.txt")
        # open(mock_file, 'w').close()
        # record_file_creation(log_dir, step_index, mock_file)
    
    sys.exit(0)

if __name__ == "__main__":
    main()
