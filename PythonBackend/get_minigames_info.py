import sys
import os
import json
import openpyxl

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
    except IndexError:
        print(json.dumps({"error": "Missing design folder path"}))
        sys.exit(1)
        
    excel_path = os.path.join(design_path, "mini_mgr.xlsx")
    
    if not os.path.exists(excel_path):
        print(json.dumps({"error": f"mini_mgr.xlsx not found at {excel_path}"}))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to load workbook: {str(e)}"}))
        sys.exit(1)

    result = {
        "max_id": 0,
        "minigames": []
    }

    ws_data = wb["活动数据"] if "活动数据" in wb.sheetnames else None
    ws_info = wb["活动信息"] if "活动信息" in wb.sheetnames else None

    if not ws_data or not ws_info:
        print(json.dumps({"error": "Missing required sheets in mini_mgr.xlsx"}))
        sys.exit(1)

    # 1. Get max ID from 活动数据
    max_id = 0
    # Search from the bottom up to find the last valid numeric ID
    for r in range(ws_data.max_row, 1, -1):
        val = ws_data.cell(row=r, column=1).value
        try:
            val_int = int(val)
            max_id = val_int
            break
        except (ValueError, TypeError):
            continue
            
    result["max_id"] = max_id

    # 2. Build history dict from 活动数据 to find 折扣链ID and 双周活动ID requirement
    history_map = {}
    
    # Columns in 活动数据 (0-indexed logic)
    # column=10 (Index 9 in python list) is 活动类型 (Minigame Name)
    # column=11 (Index 10) is 双周活动ID 
    # column=12 (Index 11) is 折扣链ID
    for r in range(ws_data.max_row, 3, -1):
        mg_name = ws_data.cell(row=r, column=10).value
        if not mg_name or not isinstance(mg_name, str):
            continue
            
        discount_id = ws_data.cell(row=r, column=12).value
        double_week_id = ws_data.cell(row=r, column=11).value
        
        if mg_name not in history_map:
            history_map[mg_name] = {
                "discount_id": discount_id if discount_id is not None else -1,
                "requires_double_week": double_week_id is not None and str(double_week_id).strip() != "None" and str(double_week_id).strip() != ""
            }

    # 3. Get all available minigames from 活动信息 (Column 2 is Alias/Name, Column 3 is internal name)
    # The requirement is that dropdown shows names from "活动信息" (user probably meant Alias, which is column 3 (index 2))
    # Wait, the user said: "available minigames from the sheet 活动信息". Let's look at row 4: [3, '大富翁', 'gembrawl', 10186, 20009]
    # So column 2 is the Chinese Name (大富翁, 答题挑战, 挖蘑菇, etc.). We should use the Chinese name because that's what's written to 活动数据.
    
    for r in range(4, ws_info.max_row + 1):
        mg_name = ws_info.cell(row=r, column=2).value
        if not mg_name or not str(mg_name).strip():
            continue
        mg_name = str(mg_name).strip()
        
        # Merge heuristic data if available
        heuristic = history_map.get(mg_name, {"discount_id": -1, "requires_double_week": False})
        
        result["minigames"].append({
            "name": mg_name,
            "discount_id": heuristic["discount_id"],
            "requires_double_week": heuristic["requires_double_week"]
        })

    print(json.dumps(result, ensure_ascii=False))

if __name__ == "__main__":
    main()
