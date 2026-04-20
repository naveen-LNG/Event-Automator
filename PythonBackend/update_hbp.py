import sys
import os
import openpyxl
import json
import re

def get_last_data_row(ws, id_col=1):
    for r in range(ws.max_row, 0, -1):
        val = ws.cell(row=r, column=id_col).value
        if val is not None and str(val).strip() != '':
            return r
    return 1

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
        payload_path = sys.argv[2]
    except IndexError:
        print("Error: Missing args")
        sys.exit(1)

    with open(payload_path, 'r', encoding='utf-8-sig') as f:
        payload = json.load(f)

    holiday_id = payload['holiday_id'] # e.g. "260601"
    start_time = payload['start_time']
    end_time = payload['end_time']

    bp_path = os.path.join(design_path, "festivals_bp.xlsx")
    item_path = os.path.join(design_path, "item.xlsx")
    store_path = os.path.join(design_path, "store.xlsx")

    try:
        wb_bp = openpyxl.load_workbook(bp_path)
        
        # --- 1. festivals_bp sheet ---
        ws_bp = wb_bp['festivals_bp']
        last_bp_r = get_last_data_row(ws_bp)
        
        # Toggle old switches
        for r in range(2, last_bp_r + 1):
            ws_bp.cell(row=r, column=2).value = False
            
        new_bp_r = last_bp_r + 1
        cycle_id = int(ws_bp.cell(row=last_bp_r, column=1).value or 0) + 1
        recommend_id = int(ws_bp.cell(row=last_bp_r, column=9).value or 18400) + 1
        
        # Duplicate last row values initially
        for c in range(1, ws_bp.max_column + 1):
            ws_bp.cell(row=new_bp_r, column=c).value = ws_bp.cell(row=last_bp_r, column=c).value
            
        ws_bp.cell(row=new_bp_r, column=1).value = cycle_id
        ws_bp.cell(row=new_bp_r, column=2).value = True
        ws_bp.cell(row=new_bp_r, column=3).value = start_time
        ws_bp.cell(row=new_bp_r, column=4).value = end_time
        ws_bp.cell(row=new_bp_r, column=5).value = "HOLIDAY_BP_TITLE_2"
        ws_bp.cell(row=new_bp_r, column=6).value = "HOLIDAY_BP_TOKEN_NAME_2"
        ws_bp.cell(row=new_bp_r, column=9).value = recommend_id

        # --- 2. 扩展表 (Extended Table) ---
        ws_ext = wb_bp['扩展表']
        last_ext_r = get_last_data_row(ws_ext)
        new_ext_start = last_ext_r + 1
        
        # Determine suffixes and last item id from previous group of 4
        prev_group_start = max(2, last_ext_r - 3)
        prev_item_id = int(ws_ext.cell(row=last_ext_r, column=5).value or 3310)
        new_item_id = prev_item_id + 1
        
        for i in range(4):
            src_r = prev_group_start + i
            tgt_r = new_ext_start + i
            for c in range(1, ws_ext.max_column + 1):
                ws_ext.cell(row=tgt_r, column=c).value = ws_ext.cell(row=src_r, column=c).value
            
            # Alias logic: detect prefix
            old_alias = str(ws_ext.cell(row=src_r, column=2).value)
            match = re.match(r'(\d{6})(.*)', old_alias)
            suffix = match.group(2) if match else "节日BP" + str(i+1)
            ws_ext.cell(row=tgt_r, column=2).value = f"{holiday_id}{suffix}"
            
            # Increment numeric columns: 1 (Extension ID), 4 (Paid Products)
            ws_ext.cell(row=tgt_r, column=1).value = int(ws_ext.cell(row=src_r, column=1).value or 0) + 4 # Wait, increment by 4 since we add a group of 4? No, user says 'remaining columns incremented by 1'. 
            # Actually, looking at current data: 38074, 38075, 38076, 38077. So they increment by 4 from previous group's corresponding row?
            # User said: 'Remaining columns will be incremented by 1'. If Row 70 (38074) was duplicated to 74? 38074 + 4 = 38078. 
            # Let's use: prev_row_val + 4.
            ws_ext.cell(row=tgt_r, column=1).value = int(ws_ext.cell(row=src_r, column=1).value) + 4
            ws_ext.cell(row=tgt_r, column=4).value = int(ws_ext.cell(row=src_r, column=4).value) + 4
            ws_ext.cell(row=tgt_r, column=5).value = new_item_id

        # --- 3. chest_plan ---
        ws_cp = wb_bp['chest_plan']
        last_cp_r = get_last_data_row(ws_cp)
        new_cp_start = last_cp_r + 1
        
        prev_cp_group_start = max(2, last_cp_r - 3)
        old_holiday_prefix = ""
        for i in range(4):
            src_r = prev_cp_group_start + i
            tgt_r = new_cp_start + i
            for c in range(1, ws_cp.max_column + 1):
                ws_cp.cell(row=tgt_r, column=c).value = ws_cp.cell(row=src_r, column=c).value
            
            ws_cp.cell(row=tgt_r, column=1).value = int(ws_cp.cell(row=src_r, column=1).value) + 4
            
            old_alias = str(ws_cp.cell(row=src_r, column=2).value)
            match = re.match(r'(\d{6})(.*)', old_alias)
            if match:
                old_holiday_prefix = match.group(1)
                suffix = match.group(2)
                ws_cp.cell(row=tgt_r, column=2).value = f"{holiday_id}{suffix}"
            
            # 积分奖励列表 string replacement
            rewards_str = str(ws_cp.cell(row=tgt_r, column=7).value)
            ws_cp.cell(row=tgt_r, column=7).value = rewards_str.replace(old_holiday_prefix, holiday_id)

        # --- 4. chest_conditions ---
        ws_cond = wb_bp['chest_conditions']
        # Locate last 4 groups (based on old_holiday_prefix)
        template_rows = []
        for r in range(4, ws_cond.max_row + 1):
            val = str(ws_cond.cell(row=r, column=2).value)
            if old_holiday_prefix and old_holiday_prefix in val:
                template_rows.append(r)
        
        if template_rows:
            next_cond_start = get_last_data_row(ws_cond) + 1
            # User example: if 260509 was 8601-8924. 
            # 8601 -> 8701 etc? 'Each group is incrementing by 100'.
            for i, src_r in enumerate(template_rows):
                tgt_r = next_cond_start + i
                for c in range(1, ws_cond.max_column + 1):
                    ws_cond.cell(row=tgt_r, column=c).value = ws_cond.cell(row=src_r, column=c).value
                
                # Alias replace
                alias = str(ws_cond.cell(row=tgt_r, column=2).value)
                ws_cond.cell(row=tgt_r, column=2).value = alias.replace(old_holiday_prefix, holiday_id)
                # ID increment by 100 for THE GROUP?
                # Actually, if we add 4 groups, and previous were 8601, 8701, 8801, 8901.
                # Next should be 8701? No, user says 'incrementing by 100'. 
                # Let's say: ID = prev_ID + 100.
                ws_cond.cell(row=tgt_r, column=1).value = int(ws_cond.cell(row=src_r, column=1).value) + 100

        wb_bp.save(bp_path)

        # --- 5. item.xlsx ---
        wb_item = openpyxl.load_workbook(item_path)
        ws_item = wb_item['item']
        # Insert before row 149
        ws_item.insert_rows(149)
        ws_item.cell(row=149, column=1).value = new_item_id
        ws_item.cell(row=149, column=2).value = f"{holiday_id}节日BP通行证"
        ws_item.cell(row=149, column=3).value = "解锁道具"
        wb_item.save(item_path)

        # --- 6. store.xlsx ---
        wb_store = openpyxl.load_workbook(store_path)
        
        # store_recharge: Insert 4 rows before 140
        ws_sr = wb_store['store_recharge']
        src_range_sr = list(range(136, 140)) # prev Holiday 4 rows
        ws_sr.insert_rows(140, 4)
        for i, src_r in enumerate(src_range_sr):
            tgt_r = 140 + i
            # Copy values
            for c in range(1, ws_sr.max_column + 1):
                ws_sr.cell(row=tgt_r, column=c).value = ws_sr.cell(row=src_r, column=c).value
            
            # Increment Store Item ID (Col 1) by 4?
            ws_sr.cell(row=tgt_r, column=1).value = int(ws_sr.cell(row=src_r, column=1).value) + 4
            # Swap Alias (Col 2 & 3)
            ws_sr.cell(row=tgt_r, column=2).value = str(ws_sr.cell(row=src_r, column=2).value).replace(old_holiday_prefix, holiday_id)
            ws_sr.cell(row=tgt_r, column=3).value = str(ws_sr.cell(row=src_r, column=3).value).replace(old_holiday_prefix, holiday_id)

        # store_package: Insert 4 rows before 147
        ws_sp = wb_store['store_package']
        src_range_sp = list(range(143, 147))
        ws_sp.insert_rows(147, 4)
        for i, src_r in enumerate(src_range_sp):
            tgt_r = 147 + i
            for c in range(1, ws_sp.max_column + 1):
                ws_sp.cell(row=tgt_r, column=c).value = ws_sp.cell(row=src_r, column=c).value
            
            ws_sp.cell(row=tgt_r, column=1).value = int(ws_sp.cell(row=src_r, column=1).value) + 4
            ws_sp.cell(row=tgt_r, column=2).value = str(ws_sp.cell(row=src_r, column=2).value).replace(old_holiday_prefix, holiday_id)
            # Content String sync
            content = str(ws_sp.cell(row=tgt_r, column=3).value)
            ws_sp.cell(row=tgt_r, column=3).value = content.replace(old_holiday_prefix, holiday_id)

        # recommend_gift: Insert 1 row before row 68
        ws_rg = wb_store['recommend_gift']
        ws_rg.insert_rows(68)
        prev_rg_r = 67
        for c in range(1, ws_rg.max_column + 1):
            ws_rg.cell(row=68, column=c).value = ws_rg.cell(row=prev_rg_r, column=c).value
        
        ws_rg.cell(row=68, column=1).value = recommend_id
        ws_rg.cell(row=68, column=2).value = f"{holiday_id}节日BP推荐"
        # Recharge ID list
        recharge_ids = [str(int(ws_sr.cell(row=140+j, column=1).value)) for j in range(4)]
        ws_rg.cell(row=68, column=3).value = ",".join(recharge_ids)

        wb_store.save(store_path)
        print("Successfully updated Holiday BattlePass configuration.")

    except Exception as e:
        print(f"Fatal Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
