import sys
import os
import json
import openpyxl

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
        payload_path = sys.argv[2]
    except IndexError:
        print("Error: Missing args")
        sys.exit(1)
        
    excel_path = os.path.join(design_path, "mini_mgr.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"Error: mini_mgr.xlsx not found at {excel_path}")
        sys.exit(1)

    try:
        with open(payload_path, 'r', encoding='utf-8-sig') as f:
            payload = json.load(f)
    except Exception as e:
        print(f"Error reading payload: {e}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Error loading workbook: {str(e)}")
        sys.exit(1)

    ws_data = wb["活动数据"] if "活动数据" in wb.sheetnames else None
    if not ws_data:
        print("Error: Missing 活动数据")
        sys.exit(1)

    TEMPLATE_IDS = {101, 102, 103}

    # Find the real last row with data (scan backwards to skip ghost empty rows)
    def get_last_data_row(ws):
        for r in range(ws.max_row, 0, -1):
            # A row is considered 'real' if column 1 (ID) has a numeric value
            try:
                val = ws.cell(row=r, column=1).value
                if val is not None and str(val).strip() != '':
                    int(val)  # must be numeric
                    return r
            except (TypeError, ValueError):
                continue
        return 1

    last_data_row = get_last_data_row(ws_data)

    # 1. Flip old switches to False — but NEVER touch template rows (101, 102, 103)
    for r in range(4, last_data_row + 1):
        row_id = ws_data.cell(row=r, column=1).value
        try:
            if int(row_id) in TEMPLATE_IDS:
                continue  # Protect template rows — always stay True
        except (TypeError, ValueError):
            continue
        if ws_data.cell(row=r, column=2).value is not None:
            ws_data.cell(row=r, column=2).value = False

    # 2. Append new rows directly after the last real data row
    start_row = last_data_row + 1
    for row_data in payload.get('rows', []):
        ws_data.cell(row=start_row, column=1).value = row_data['id']
        ws_data.cell(row=start_row, column=2).value = True
        
        # Start time
        ws_data.cell(row=start_row, column=5).value = row_data['start_time']
        # End time
        ws_data.cell(row=start_row, column=6).value = row_data['end_time']
        
        # Minigame
        ws_data.cell(row=start_row, column=10).value = row_data['minigame']
        
        # Double Week ID
        dw_val = row_data.get('double_week_id')
        if dw_val and str(dw_val).strip() != "":
            try:
                ws_data.cell(row=start_row, column=11).value = int(dw_val)
            except ValueError:
                ws_data.cell(row=start_row, column=11).value = dw_val
                
        # Discount ID
        if row_data.get('discount_id') != -1:
            ws_data.cell(row=start_row, column=12).value = row_data['discount_id']
        
        start_row += 1

    try:
        wb.save(excel_path)
    except PermissionError:
        print(f"ERROR: Cannot save '{excel_path}'.\nThe file is open in Excel or another program. Please close it and try again.")
        sys.exit(1)
    print(f"Successfully appended {len(payload.get('rows', []))} Minigame rows to Excel.")

if __name__ == "__main__":
    main()
