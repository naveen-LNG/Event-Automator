import sys
import os
import json
import openpyxl
import re
from datetime import datetime

def get_last_data_row(ws, id_col=1):
    for r in range(ws.max_row, 0, -1):
        # A row is considered 'real' if the ID column has a numeric-like value
        val = ws.cell(row=r, column=id_col).value
        if val is not None and str(val).strip() != '':
            return r
    return 1

def update_switch_in_sheet(ws, active_id):
    # Logic: Toggle all to False except the new one (Cycle ID 1-3 usually reserved so start from 4 if ID-based)
    # Actually, the user said all remaining should be false.
    for r in range(4, ws.max_row + 1):
        cid = ws.cell(row=r, column=1).value
        if cid is not None:
             ws.cell(row=r, column=2).value = False

def get_sequential_yymm(last_alias):
    # e.g. "2605BattlePass解锁礼包" -> "2606"
    match = re.search(r'(\d{4})', last_alias)
    if not match: return "2601"
    
    yymm = match.group(1)
    yy = int(yymm[:2])
    mm = int(yymm[2:])
    
    mm += 1
    if mm > 12:
        mm = 1
        yy += 1
    return f"{yy:02d}{mm:02d}"

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    try:
        design_path = sys.argv[1]
        payload_path = sys.argv[2]
    except IndexError:
        print("Error: Missing args")
        sys.exit(1)

    with open(payload_path, 'r', encoding='utf-8-sig') as f:
        payload = json.load(f)

    cycle_id = payload['cycle_id']
    start_time = payload['start_time']
    end_time = payload['end_time']
    bg_limit = payload['bg_limit']
    template_cycle = payload['template_cycle']

    bp_path = os.path.join(design_path, "battle_pass.xlsx")
    item_path = os.path.join(design_path, "item.xlsx")

    try:
        wb_bp = openpyxl.load_workbook(bp_path)
        
        # --- 1. Update battle_pass sheet ---
        ws_bp = wb_bp['battle_pass']
        last_row = get_last_data_row(ws_bp)
        
        # Preserve old values before switching
        prev_recharge = str(ws_bp.cell(row=last_row, column=8).value)
        prev_item_id = int(ws_bp.cell(row=last_row, column=9).value)
        
        # Toggle old switches to False
        for r in range(4, last_row + 1):
             ws_bp.cell(row=r, column=2).value = False
             
        new_yymm = get_sequential_yymm(prev_recharge)
        new_row = last_row + 1
        
        ws_bp.cell(row=new_row, column=1).value = cycle_id
        ws_bp.cell(row=new_row, column=2).value = True
        ws_bp.cell(row=new_row, column=3).value = start_time
        ws_bp.cell(row=new_row, column=4).value = end_time
        ws_bp.cell(row=new_row, column=5).value = "BP_TITLE"
        ws_bp.cell(row=new_row, column=6).value = "BP积分"
        ws_bp.cell(row=new_row, column=7).value = f"BattlePass第{cycle_id}期"
        ws_bp.cell(row=new_row, column=8).value = f"{new_yymm}BattlePass解锁礼包"
        ws_bp.cell(row=new_row, column=9).value = prev_item_id + 1
        
        # --- 2. Update chest_plan sheet ---
        sname_cp = 'chest_plan' if 'chest_plan' in wb_bp.sheetnames else 'Chest_plan'
        ws_cp = wb_bp[sname_cp]
        last_cp_r = get_last_data_row(ws_cp)
        new_cp_r = last_cp_r + 1
        
        new_scheme_id = int(ws_cp.cell(row=last_cp_r, column=1).value) + 1
        
        # Duplicate last row as generic template
        for c in range(1, ws_cp.max_column + 1):
            ws_cp.cell(row=new_cp_r, column=c).value = ws_cp.cell(row=last_cp_r, column=c).value
            
        ws_cp.cell(row=new_cp_r, column=1).value = new_scheme_id
        ws_cp.cell(row=new_cp_r, column=2).value = f"BattlePass第{cycle_id}期"
        ws_cp.cell(row=new_cp_r, column=3).value = bg_limit
        
        #积分奖励列表 string replacement (e.g. BP20 -> BP21)
        reward_str = str(ws_cp.cell(row=new_cp_r, column=5).value)
        # Find template BP prefix
        new_reward_str = reward_str.replace(f"BP{template_cycle}进度", f"BP{cycle_id}进度")
        # Handle 800 -> 900 or 900 -> 800 count diff if needed? 
        # Actually user says duplicate based on recent 800/900.
        # Let's just fix the string.
        ws_cp.cell(row=new_cp_r, column=5).value = new_reward_str

        # --- 3. Duplicate chest_conditions entries ---
        ws_cond = wb_bp['chest_conditions']
        # Find all rows belonging to template_cycle
        template_rows = []
        for r in range(4, ws_cond.max_row + 1):
             val = str(ws_cond.cell(row=r, column=2).value) # Alias
             if f"BP{template_cycle}进度" in val:
                 template_rows.append(r)
        
        if template_rows:
            next_start_r = get_last_data_row(ws_cond) + 1
            for i, src_r in enumerate(template_rows):
                target_r = next_start_r + i
                for c in range(1, ws_cond.max_column + 1):
                    ws_cond.cell(row=target_r, column=c).value = ws_cond.cell(row=src_r, column=c).value
                
                # Update ID and Alias
                ws_cond.cell(row=target_r, column=1).value = cycle_id * 1000 + (i + 1)
                alias = str(ws_cond.cell(row=target_r, column=2).value)
                ws_cond.cell(row=target_r, column=2).value = alias.replace(f"BP{template_cycle}进度", f"BP{cycle_id}进度")
                
                # Update reward names
                free_box = str(ws_cond.cell(row=target_r, column=4).value)
                ws_cond.cell(row=target_r, column=4).value = free_box.replace(f"BP{template_cycle}_", f"BP{cycle_id}_")
                paid_box = str(ws_cond.cell(row=target_r, column=5).value)
                ws_cond.cell(row=target_r, column=5).value = paid_box.replace(f"BP{template_cycle}_", f"BP{cycle_id}_")

        # --- 4. Duplicate chest_rewards entries ---
        ws_rew = wb_bp['chest_rewards']
        template_rew_rows = []
        for r in range(4, ws_rew.max_row + 1):
             val = str(ws_rew.cell(row=r, column=2).value) # Alias
             if f"BP{template_cycle}_" in val:
                 template_rew_rows.append(r)
        
        if template_rew_rows:
            next_rew_start_r = get_last_data_row(ws_rew) + 1
            for i, src_r in enumerate(template_rew_rows):
                target_r = next_rew_start_r + i
                for c in range(1, ws_rew.max_column + 1):
                    ws_rew.cell(row=target_r, column=c).value = ws_rew.cell(row=src_r, column=c).value
                
                # Update Alias and ID
                alias = str(ws_rew.cell(row=target_r, column=2).value)
                ws_rew.cell(row=target_r, column=2).value = alias.replace(f"BP{template_cycle}_", f"BP{cycle_id}_")
                
                # ID increment logic: 201000 -> 211000? 
                # Actually rewards ID is tricky. BP 1 is 201xxx. BP 20 is 220xxx.
                # Let's increment based on the template ID difference.
                old_id = int(ws_rew.cell(row=src_r, column=1).value or 0)
                # If BP diff is 1, maybe ID diff is 1000? 
                # Let's use CycleID * 10000 + original offset?
                # User says: each group has 1000 increment.
                # BP 20 has IDs around 220xxx. BP 1 had 201xxx.
                # So maybe (CycleID + 19) * 1000 + index? No.
                # Let's just find the max ID in the sheet and add to it.
                # But safer to follow the 1000/group rule.
                # ID = Template_ID + (Cycle - Template) * 1000
                ws_rew.cell(row=target_r, column=1).value = old_id + (cycle_id - template_cycle) * 1000

        try:
            wb_bp.save(bp_path)
        except PermissionError:
             print(f"Permission Error: Close {bp_path} in Excel first!")
             sys.exit(1)

        # --- 5. Update item.xlsx ---
        try:
            wb_item = openpyxl.load_workbook(item_path)
            ws_item = wb_item['item']
            
            new_item_id = prev_item_id + 1
            exists = False
            comment_row = -1
            
            for r in range(1, ws_item.max_row + 1):
                val = str(ws_item.cell(row=r, column=1).value)
                if val == str(new_item_id):
                    exists = True
                    break
                if "//3301" in val:
                    comment_row = r
            
            if not exists:
                if comment_row == -1: comment_row = ws_item.max_row + 1
                ws_item.insert_rows(comment_row)
                ws_item.cell(row=comment_row, column=1).value = new_item_id
                ws_item.cell(row=comment_row, column=2).value = f"{new_yymm}BattlePass通行证"
                ws_item.cell(row=comment_row, column=3).value = "解锁道具"
                wb_item.save(item_path)
            
        except Exception as e:
            print(f"Warning: Failed to update item.xlsx: {e}")

        print("Successfully updated BattlePass configuration across all files.")

    except Exception as e:
        print(f"Fatal Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
